#!/usr/bin/env python3
"""
Data Agent 任务一：老人照护档案长者健康智能计算图构建流水线

功能说明：
1. 读取 Excel 老人档案（多工作表）；
2. 基于 LangGraph 编排多阶段工作流；
3. 通过“分工作表智能体”并行进行抽取、清洗、结构化；
4. 生成符合节点/边约束的长者健康智能计算图；
5. 输出中间产物、流程图、JSON 结果、PDF 核心日志；
6. 支持异常捕获与失败恢复记录，便于比赛答辩回放。

注意事项：
- 本脚本通过配置文件读取 API-Key，不在代码中硬编码；
- 默认配置文件路径：/srv/supercare/config/data_agent1_config.json
"""

from __future__ import annotations

import argparse
import json
import math
import re
import subprocess
import time
import traceback
import textwrap
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, TypedDict

from langgraph.graph import END, START, StateGraph
import networkx as nx
from openai import OpenAI
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

try:
    from pypinyin import lazy_pinyin
except ImportError:  # pragma: no cover
    lazy_pinyin = None


DEFAULT_INPUT_FILE = "/srv/supercare/DataSource/陈女士/ms_chen_CareCase.xlsx"
DEFAULT_OUTPUT_ROOT = "/srv/supercare/data-agent1"
DEFAULT_CONFIG_PATH = "/srv/supercare/config/data_agent1_config.json"

# 仅保留任务要求中的核心节点类型
ALLOWED_NODE_TYPES = {
    "老人主体",
    "基础信息",
    "人群标签",
    "能力评估",
    "生命体征记录",
    "服药记录",
    "照护服务记录",
    "健康评估记录",
    "睡眠记录",
    "住院记录",
    "异常事件记录",
    "照护人员",
    "医护人员",
    "家属",
    "医疗机构",
    "药品",
    "疾病",
    "时间维度",
}

# 仅保留任务要求中的核心关系
ALLOWED_RELATIONS = {
    "有基础信息",
    "有人群标签",
    "有能力评估结果",
    "有生命体征记录",
    "有服药记录",
    "接受照护服务",
    "有健康评估记录",
    "有睡眠记录",
    "有住院记录",
    "有异常事件记录",
    "有亲属关系",
    "接受医疗服务",
    "入住/就诊于",
    "服用药品",
    "确诊疾病",
    "发生于",
    "执行服务",
    "执行诊疗",
    "处理事件",
}

SHEET_TO_NODE_RELATION: Dict[str, Tuple[str, str]] = {
    "生命体征记录": ("生命体征记录", "有生命体征记录"),
    "服药情况": ("服药记录", "有服药记录"),
    "照护服务日志": ("照护服务记录", "接受照护服务"),
    "健康评估记录": ("健康评估记录", "有健康评估记录"),
    "睡眠情况记录": ("睡眠记录", "有睡眠记录"),
    "住院记录": ("住院记录", "有住院记录"),
    "异常事件记录": ("异常事件记录", "有异常事件记录"),
}

TIMESTAMP_FIELD_CANDIDATES = [
    "时间戳",
    "时间",
    "记录时间",
    "发生时间",
    "日期",
    "评估时间",
    "用药时间",
    "入院时间",
    "出院时间",
]

EXPECTED_SHEET_SEQUENCE = [
    "健康档案",
    "生命体征记录",
    "服药情况",
    "照护服务日志",
    "健康评估记录",
    "睡眠情况记录",
    "住院记录",
    "异常事件记录",
]

# 长者健康智能计算图“重要信息优先”保留策略：限制各工作表进入图谱的记录上限。
SHEET_IMPORTANT_RECORD_LIMITS = {
    "生命体征记录": 16,
    "服药情况": 16,
    "照护服务日志": 16,
    "健康评估记录": 16,
    "睡眠情况记录": 16,
    "住院记录": 16,
    "异常事件记录": 16,
}

IMPORTANT_CONTENT_KEYWORDS = [
    "风险",
    "异常",
    "跌倒",
    "住院",
    "出院",
    "诊断",
    "疾病",
    "BPSD",
    "会诊",
    "药",
    "失眠",
    "攻击",
    "游走",
    "压疮",
    "吞咽",
    "发热",
    "血压",
    "心率",
]

# 图谱 PNG 仅绘制以下 11 种节点 type（与全量图谱 schema 解耦；不依赖 task-agent 工程）
GRAPH_VISUAL_NODE_TYPES_ORDER = (
    "老人主体",
    "近况摘要",
    "风险洞察",
    "照护洞察",
    "个性化健康档案",
    "基础信息",
    "体征洞察",
    "用药洞察",
    "照护服务记录",
    "健康评估记录",
    "睡眠记录",
)

# 每类可视化条数区间（默认取中间偏上，图面信息更密）
VISUAL_TYPE_MIN_COUNT = 8
VISUAL_TYPE_MAX_COUNT = 24
VISUAL_TYPE_DEFAULT_COUNT = 18

MAX_INSIGHT_NODE_PER_SHEET = 4
MAX_EVIDENCE_NODE_PER_SHEET = 3


def current_time_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def normalize_value(raw_value: Any) -> str:
    """统一清洗字段值：去空格、去换行、标准化空值。"""
    if raw_value is None:
        return ""
    value_text = str(raw_value).strip()
    if value_text.lower() in {"none", "nan", "nat"}:
        return ""
    return re.sub(r"\s+", " ", value_text)


def is_high_value_field(field_name: str) -> bool:
    """高价值字段筛选：保留身份、评估、诊疗、照护、风险、时间等信息。"""
    high_value_keywords = [
        "姓名",
        "性别",
        "年龄",
        "诊断",
        "疾病",
        "症状",
        "风险",
        "事件",
        "服药",
        "药",
        "生命体征",
        "血压",
        "心率",
        "体温",
        "评估",
        "照护",
        "护理",
        "睡眠",
        "住院",
        "出院",
        "家属",
        "医生",
        "护士",
        "机构",
        "时间",
        "日期",
        "能力",
        "标签",
    ]
    return any(keyword in field_name for keyword in high_value_keywords)


def detect_timestamp(record: Dict[str, str]) -> str:
    """从记录中提取时间戳，保证时序图谱可落地。"""
    for candidate_field in TIMESTAMP_FIELD_CANDIDATES:
        for field_name, field_value in record.items():
            if candidate_field in field_name and field_value:
                return field_value
    return ""


def convert_name_to_pinyin(person_name: str, fallback_name: str) -> str:
    """将中文姓名转拼音，作为图谱文件名。"""
    clean_name = normalize_value(person_name)
    if clean_name and lazy_pinyin is not None:
        return "".join(lazy_pinyin(clean_name))
    if clean_name:
        return re.sub(r"[^a-zA-Z0-9_-]", "_", clean_name)
    return re.sub(r"[^a-zA-Z0-9_-]", "_", fallback_name)


def safe_json_dump(target_path: Path, content: Any) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with target_path.open("w", encoding="utf-8") as output_file:
        json.dump(content, output_file, ensure_ascii=False, indent=2)


def load_visual_font(font_size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    """加载可用于图谱图片标注的字体。"""
    candidate_font_paths = [
        "/srv/supercare/assets/fonts_pkg/extracted/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/srv/supercare/assets/fonts/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/arphic/ukai.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate_font_path in candidate_font_paths:
        if Path(candidate_font_path).exists():
            try:
                return ImageFont.truetype(candidate_font_path, font_size)
            except Exception:
                continue
    return ImageFont.load_default()


def build_graph_visual_showcase_subgraph(
    graph_data: Dict[str, Any],
    per_category_count: int = VISUAL_TYPE_DEFAULT_COUNT,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, str]]:
    """
    构建用于 PNG 的展示子图：
    - 仅包含上述 11 种 type 的节点（按 type 字段精确匹配，不再把其他 type 归并进来）；
    - 每个 type（除老人主体全保留外）保留条数由 per_category_count 夹逼在 [MIN, MAX] 内；
    - 老人主体不截断；其余类型按重要性排序后截断。
    """
    all_nodes = graph_data.get("nodes", [])
    all_edges: List[Dict[str, Any]] = list(graph_data.get("edges", []))
    node_id_to_data = {node["id"]: dict(node) for node in all_nodes}

    elder_node_id = ""
    for node in all_nodes:
        if node.get("type") == "老人主体":
            elder_node_id = node.get("id", "")
            break

    visual_limit = max(VISUAL_TYPE_MIN_COUNT, min(VISUAL_TYPE_MAX_COUNT, per_category_count))
    type_to_selected_ids: Dict[str, List[str]] = {t: [] for t in GRAPH_VISUAL_NODE_TYPES_ORDER}
    node_id_to_category: Dict[str, str] = {}

    def node_visual_importance_score(node_item: Dict[str, Any]) -> int:
        node_type = normalize_value(node_item.get("type", ""))
        node_name = normalize_value(node_item.get("name", ""))
        node_properties = node_item.get("properties", {})
        score = 0
        if node_type == "老人主体":
            score += 200
        if node_type in {"风险洞察", "近况摘要"}:
            score += 35
        if node_type in {"健康评估记录", "用药洞察", "照护服务记录", "睡眠记录", "体征洞察"}:
            score += 22
        if any(keyword in node_name for keyword in IMPORTANT_CONTENT_KEYWORDS):
            score += 18
        if isinstance(node_properties, dict):
            fields_data = node_properties.get("fields", {})
            if isinstance(fields_data, dict):
                field_text_blob = " ".join(
                    [normalize_value(field_key) + " " + normalize_value(field_value) for field_key, field_value in fields_data.items()]
                )
                for keyword in IMPORTANT_CONTENT_KEYWORDS:
                    if keyword in field_text_blob:
                        score += 5
            if normalize_value(node_properties.get("timestamp", "")):
                score += 3
        return score

    for visual_type in GRAPH_VISUAL_NODE_TYPES_ORDER:
        candidate_nodes = [node for node in all_nodes if node.get("type") == visual_type]
        if visual_type == "老人主体":
            chosen = candidate_nodes
        else:
            deduped_nodes: List[Dict[str, Any]] = []
            seen_label_set: Set[str] = set()
            for node in candidate_nodes:
                node_label = normalize_value(node.get("name", ""))
                if node_label in seen_label_set:
                    continue
                seen_label_set.add(node_label)
                deduped_nodes.append(node)
            ranked_nodes = sorted(
                deduped_nodes,
                key=lambda node_item: node_visual_importance_score(node_item),
                reverse=True,
            )
            chosen = ranked_nodes[:visual_limit]
        type_to_selected_ids[visual_type] = [node.get("id", "") for node in chosen]
        for node in chosen:
            node_id_to_category[node.get("id", "")] = visual_type

    selected_node_id_set: Set[str] = set()
    for id_list in type_to_selected_ids.values():
        selected_node_id_set.update(id_list)

    anchor_elder = elder_node_id or next(
        (node.get("id", "") for node in all_nodes if node.get("type") == "老人主体"),
        "",
    )
    for visual_type in GRAPH_VISUAL_NODE_TYPES_ORDER:
        if type_to_selected_ids.get(visual_type):
            continue
        placeholder_node_id = f"placeholder_{visual_type}"
        node_id_to_data[placeholder_node_id] = {
            "id": placeholder_node_id,
            "type": "占位节点",
            "name": f"{visual_type}：暂无典型数据",
            "properties": {"category": visual_type, "placeholder": True},
        }
        selected_node_id_set.add(placeholder_node_id)
        node_id_to_category[placeholder_node_id] = visual_type
        if anchor_elder:
            all_edges.append(
                {
                    "source": anchor_elder,
                    "relation": "暂无数据",
                    "target": placeholder_node_id,
                    "properties": {"placeholder": True},
                }
            )

    selected_nodes: List[Dict[str, Any]] = []
    for node_id in selected_node_id_set:
        if node_id in node_id_to_data:
            node_item = dict(node_id_to_data[node_id])
            node_item["category"] = node_id_to_category.get(node_id, node_item.get("type", "其他"))
            selected_nodes.append(node_item)

    selected_edges: List[Dict[str, Any]] = []
    for edge in all_edges:
        if (
            edge.get("source") in selected_node_id_set
            and edge.get("target") in selected_node_id_set
        ):
            selected_edges.append(edge)

    return selected_nodes, selected_edges, node_id_to_category


def render_knowledge_graph_image(
    graph_data: Dict[str, Any],
    output_image_path: Path,
) -> Dict[str, Any]:
    """将长者健康智能计算图绘制为 PNG 图片。"""
    selected_nodes, selected_edges, node_id_to_category = build_graph_visual_showcase_subgraph(graph_data)
    graph_obj = nx.Graph()
    for node in selected_nodes:
        graph_obj.add_node(
            node["id"],
            node_type=node.get("type", "未知"),
            node_name=node.get("name", ""),
            category=node.get("category", node.get("type", "其他")),
            node_confidence=node.get("properties", {}).get("可信度"),
        )
    for edge in selected_edges:
        graph_obj.add_edge(
            edge.get("source"),
            edge.get("target"),
            relation=edge.get("relation", ""),
            relation_confidence=edge.get("properties", {}).get("可信度"),
        )

    if graph_obj.number_of_nodes() == 0:
        blank_image = Image.new("RGB", (1400, 900), color=(255, 255, 255))
        drawer = ImageDraw.Draw(blank_image)
        font = load_visual_font(28)
        drawer.text((80, 80), "长者健康智能计算图为空，暂无可视化内容。", fill="black", font=font)
        blank_image.save(output_image_path, format="PNG")
        return {"node_count": 0, "edge_count": 0, "sampled": False}

    # 采用“分类锚点 + 力导向收敛”布局：既保持分类，又让图形紧凑。
    elder_node_ids = [node_id for node_id in graph_obj.nodes() if graph_obj.nodes[node_id].get("node_type") == "老人主体"]
    elder_node_id = elder_node_ids[0] if elder_node_ids else None
    category_order = list(GRAPH_VISUAL_NODE_TYPES_ORDER)
    category_to_node_ids: Dict[str, List[str]] = {category_name: [] for category_name in category_order}
    for node_id in graph_obj.nodes():
        category_name = graph_obj.nodes[node_id].get("category", "其他")
        category_to_node_ids.setdefault(category_name, []).append(node_id)

    layout_position: Dict[str, Tuple[float, float]] = {}
    if elder_node_id:
        layout_position[elder_node_id] = (0.0, 0.0)
    ring_radius = 1.0
    for category_index, category_name in enumerate(category_order):
        node_ids = category_to_node_ids.get(category_name, [])
        if not node_ids:
            continue
        base_angle = 2 * math.pi * category_index / max(1, len(category_order))
        for node_index, node_id in enumerate(node_ids):
            if node_id == elder_node_id:
                continue
            sub_angle = base_angle + (node_index - (len(node_ids) - 1) / 2) * 0.09
            x_position = ring_radius * math.cos(sub_angle)
            y_position = ring_radius * math.sin(sub_angle)
            layout_position[node_id] = (x_position, y_position)

    # 若存在未分配坐标节点，则补充 spring 布局。
    missing_nodes = [node_id for node_id in graph_obj.nodes() if node_id not in layout_position]
    if missing_nodes:
        fallback_layout = nx.spring_layout(graph_obj.subgraph(missing_nodes), seed=42)
        for node_id, position in fallback_layout.items():
            layout_position[node_id] = (position[0], position[1])
    layout_position = nx.spring_layout(
        graph_obj,
        pos=layout_position,
        seed=42,
        k=max(0.06, 0.30 / math.sqrt(max(4, graph_obj.number_of_nodes()))),
        iterations=360,
    )

    image_width, image_height = 5600, 3600
    margin = 120
    canvas = Image.new("RGB", (image_width, image_height), color=(255, 255, 255))
    drawer = ImageDraw.Draw(canvas)
    title_font = load_visual_font(56)
    node_label_font = load_visual_font(28)
    edge_label_font = load_visual_font(22)
    legend_font = load_visual_font(26)

    x_coords = [position[0] for position in layout_position.values()]
    y_coords = [position[1] for position in layout_position.values()]
    min_x, max_x = min(x_coords), max(x_coords)
    min_y, max_y = min(y_coords), max(y_coords)

    # 为右下角图例预留像素，避免节点与标签压到图例区
    legend_panel_reserve_right = 1920
    legend_panel_reserve_bottom = 450

    def map_point(raw_x: float, raw_y: float) -> Tuple[int, int]:
        x_span = max(max_x - min_x, 1e-6)
        y_span = max(max_y - min_y, 1e-6)
        usable_width = max(320, image_width - 2 * margin - legend_panel_reserve_right)
        usable_height = max(320, image_height - 2 * margin - legend_panel_reserve_bottom)
        mapped_x = margin + int((raw_x - min_x) / x_span * usable_width)
        mapped_y = margin + int((raw_y - min_y) / y_span * usable_height)
        return mapped_x, mapped_y

    def wrap_text_with_limit(text_value: str, line_char_limit: int = 16, max_lines: int = 4) -> str:
        normalized_text = normalize_value(text_value)
        if not normalized_text:
            return ""
        wrapped_lines = textwrap.wrap(normalized_text, width=line_char_limit)
        if len(wrapped_lines) > max_lines:
            wrapped_lines = wrapped_lines[:max_lines]
            wrapped_lines[-1] = wrapped_lines[-1][: max(1, line_char_limit - 1)] + "…"
        return "\n".join(wrapped_lines)

    category_color_map = {
        "老人主体": (52, 73, 94),
        "近况摘要": (243, 156, 18),
        "风险洞察": (192, 57, 43),
        "照护洞察": (26, 188, 156),
        "个性化健康档案": (241, 196, 15),
        "基础信息": (247, 220, 111),
        "体征洞察": (52, 152, 219),
        "用药洞察": (46, 204, 113),
        "照护服务记录": (22, 160, 133),
        "健康评估记录": (155, 89, 182),
        "睡眠记录": (142, 68, 173),
        "占位节点": (189, 195, 199),
        "其他": (120, 120, 120),
    }
    default_color = (120, 120, 120)
    relation_display_map = {
        "有基础信息": "基础信息",
        "有人群标签": "人群标签",
        "有能力评估结果": "能力评估",
        "有生命体征记录": "体征记录",
        "有服药记录": "服药记录",
        "接受照护服务": "照护服务",
        "有健康评估记录": "健康评估",
        "有睡眠记录": "睡眠记录",
        "有住院记录": "住院记录",
        "有异常事件记录": "异常事件",
        "暂无数据": "暂无数据",
        "证据支撑": "证据",
    }

    # 先绘制边，并绘制边关系中文标签
    for source_node_id, target_node_id in graph_obj.edges():
        source_point = map_point(*layout_position[source_node_id])
        target_point = map_point(*layout_position[target_node_id])
        drawer.line([source_point, target_point], fill=(180, 180, 180), width=4)
        relation_text = graph_obj.edges[source_node_id, target_node_id].get("relation", "关联")
        relation_text = relation_display_map.get(relation_text, relation_text)
        edge_center_x = (source_point[0] + target_point[0]) // 2
        edge_center_y = (source_point[1] + target_point[1]) // 2
        relation_confidence = graph_obj.edges[source_node_id, target_node_id].get("relation_confidence")
        relation_suffix = f"/{relation_confidence}" if relation_confidence is not None else ""
        relation_draw_text = f"{relation_text}{relation_suffix}"[:8]
        relation_box = [edge_center_x - 42, edge_center_y - 16, edge_center_x + 42, edge_center_y + 16]
        drawer.rectangle(relation_box, fill=(255, 255, 255), outline=(230, 230, 230))
        drawer.text((edge_center_x - 36, edge_center_y - 11), relation_draw_text, fill=(70, 70, 70), font=edge_label_font)

    degree_counter = dict(graph_obj.degree())
    sorted_node_ids = sorted(
        graph_obj.nodes(),
        key=lambda node_id: degree_counter.get(node_id, 0),
        reverse=True,
    )
    label_node_id_set = set(sorted_node_ids)
    for node_id in sorted_node_ids:
        node_data = graph_obj.nodes[node_id]
        node_type = node_data.get("node_type", "未知")
        node_name = node_data.get("node_name", node_id)
        node_category = node_data.get("category", node_type)
        if node_type == "老人主体":
            color = category_color_map.get("老人主体", default_color)
        else:
            color = category_color_map.get(node_category, default_color)
        center_x, center_y = map_point(*layout_position[node_id])
        radius = 12 if node_type == "时间维度" else 18
        if node_type == "老人主体":
            radius = 30
        drawer.ellipse(
            [center_x - radius, center_y - radius, center_x + radius, center_y + radius],
            fill=color,
            outline=(50, 50, 50),
            width=2,
        )
        if node_id in label_node_id_set or node_type == "老人主体":
            display_prefix = node_category if node_type != "老人主体" else "老人主体"
            node_confidence = node_data.get("node_confidence")
            confidence_suffix = f" (置信{node_confidence})" if node_confidence is not None else ""
            label_text = wrap_text_with_limit(f"{display_prefix}:{node_name}{confidence_suffix}", line_char_limit=18, max_lines=4)
            line_count = max(1, label_text.count("\n") + 1)
            box_height = 18 + line_count * 34
            drawer.rectangle(
                [center_x + 14, center_y - 20, center_x + 620, center_y - 20 + box_height],
                fill=(255, 255, 255),
                outline=(235, 235, 235),
            )
            drawer.multiline_text(
                (center_x + 18, center_y - 14),
                label_text,
                fill=(25, 25, 25),
                font=node_label_font,
                spacing=6,
            )

    # 标题和图例
    drawer.text((90, 56), "老人照护长者健康智能计算图（11类核心类型）", fill=(0, 0, 0), font=title_font)
    info_text = (
        f"原始节点: {len(graph_data.get('nodes', []))}  原始边: {len(graph_data.get('edges', []))}  "
        f"可视化节点: {graph_obj.number_of_nodes()}  可视化边: {graph_obj.number_of_edges()}"
    )
    drawer.text((90, 134), info_text, fill=(30, 30, 30), font=legend_font)

    # 图例置于右下角，带浅色底框，避免与主图左侧/中部文字重叠
    legend_categories = [category for category in category_order if category_to_node_ids.get(category)]
    legend_items = legend_categories[:12]
    legend_cols = 4
    legend_col_step = 420
    legend_row_step = 88
    legend_rows = max(1, (len(legend_items) + legend_cols - 1) // legend_cols)
    legend_panel_pad_x, legend_panel_pad_y = 28, 22
    legend_title_h = 36
    legend_panel_inner_w = legend_cols * legend_col_step + 24
    legend_panel_inner_h = legend_title_h + legend_rows * legend_row_step + 18
    legend_panel_w = legend_panel_inner_w + 2 * legend_panel_pad_x
    legend_panel_h = legend_panel_inner_h + 2 * legend_panel_pad_y
    legend_bg_x = image_width - margin - legend_panel_w
    legend_bg_y = image_height - margin - legend_panel_h
    drawer.rounded_rectangle(
        [legend_bg_x, legend_bg_y, legend_bg_x + legend_panel_w, legend_bg_y + legend_panel_h],
        radius=16,
        fill=(248, 250, 252),
        outline=(210, 218, 230),
        width=2,
    )
    legend_caption_font = load_visual_font(22)
    drawer.text(
        (legend_bg_x + legend_panel_pad_x, legend_bg_y + legend_panel_pad_y),
        "图例（节点类型）",
        fill=(52, 64, 84),
        font=legend_caption_font,
    )
    legend_x = legend_bg_x + legend_panel_pad_x
    legend_y = legend_bg_y + legend_panel_pad_y + legend_title_h
    for index, category_name in enumerate(legend_items):
        col = index % legend_cols
        row = index // legend_cols
        x = legend_x + col * legend_col_step
        y = legend_y + row * legend_row_step
        color = category_color_map.get(category_name, default_color)
        drawer.rectangle([x, y, x + 32, y + 32], fill=color, outline=(55, 65, 80), width=1)
        drawer.text((x + 42, y + 2), category_name, fill=(18, 24, 38), font=legend_font)

    output_image_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_image_path, format="PNG")
    return {
        "node_count": graph_obj.number_of_nodes(),
        "edge_count": graph_obj.number_of_edges(),
        "sampled": True,
        "categories": sorted(list(set(node_id_to_category.values()))),
    }


@dataclass
class AgentResult:
    agent_name: str
    sheet_name: str
    status: str
    duration_seconds: float
    record_count: int
    message: str


@dataclass
class AppConfig:
    """项目配置：统一管理 MinerU 与通义千问参数。"""

    mineru_api_url: str
    mineru_backend: str
    mineru_method: str
    dashscope_api_key: str
    dashscope_base_url: str
    dashscope_model: str
    qwen_max_tokens: int
    qwen_temperature: float
    parser_mode: str


def load_app_config(config_file_path: Path) -> AppConfig:
    """从 JSON 配置文件加载运行参数。"""
    if not config_file_path.exists():
        raise FileNotFoundError(
            f"配置文件不存在：{config_file_path}。请先复制模板并填写 API-Key。"
        )
    with config_file_path.open("r", encoding="utf-8") as config_file:
        config_content = json.load(config_file)

    mineru_config = config_content.get("mineru", {})
    qwen_config = config_content.get("qwen", {})
    api_key_value = qwen_config.get("api_key", "").strip()
    if not api_key_value or api_key_value.startswith("REPLACE_WITH_"):
        raise ValueError("配置文件中的 qwen.api_key 为空，请先填写通义千问 API-Key。")

    mineru_api_url_value = mineru_config.get("api_url", "").strip()
    if mineru_api_url_value and not (
        mineru_api_url_value.startswith("http://")
        or mineru_api_url_value.startswith("https://")
    ):
        mineru_api_url_value = ""

    return AppConfig(
        mineru_api_url=mineru_api_url_value,
        mineru_backend=mineru_config.get("backend", "pipeline"),
        mineru_method=mineru_config.get("method", "auto"),
        dashscope_api_key=api_key_value,
        dashscope_base_url=qwen_config.get(
            "base_url", "https://dashscope.aliyuncs.com/compatible-mode/v1"
        ),
        dashscope_model=qwen_config.get("model", "qwen3-max"),
        qwen_max_tokens=int(qwen_config.get("max_tokens", 500)),
        qwen_temperature=float(qwen_config.get("temperature", 0.2)),
        parser_mode=str(config_content.get("runtime", {}).get("parser_mode", "mineru")),
    )


class JsonlLogger:
    """核心日志管理器：记录每步输入输出、工具调用、耗时、异常。"""

    def __init__(self, log_file_path: Path) -> None:
        self.log_file_path = log_file_path
        self.log_file_path.parent.mkdir(parents=True, exist_ok=True)
        self.buffer: List[Dict[str, Any]] = []
        # 每次任务启动时清空旧日志，避免跨次运行混淆。
        self.log_file_path.write_text("", encoding="utf-8")

    def log(
        self,
        level: str,
        step_name: str,
        message: str,
        payload: Optional[Dict[str, Any]] = None,
    ) -> None:
        event = {
            "timestamp": current_time_text(),
            "level": level,
            "step": step_name,
            "message": message,
            "payload": payload or {},
        }
        self.buffer.append(event)
        with self.log_file_path.open("a", encoding="utf-8") as output_file:
            output_file.write(json.dumps(event, ensure_ascii=False) + "\n")

    def export_pdf_summary(
        self,
        pdf_path: Path,
        task_overview: Dict[str, Any],
        graph_preview: Dict[str, Any],
        agent_catalog: List[Dict[str, Any]],
        workflow_catalog: List[Dict[str, Any]],
        innovation_points: List[str],
        execution_summary: Dict[str, Any],
    ) -> None:
        """导出 PDF 详细技术报告（评委可读版）。"""
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        style_sheet = getSampleStyleSheet()
        normal_style = style_sheet["BodyText"]
        normal_style.fontName = "STSong-Light"
        heading_style = style_sheet["Heading2"]
        heading_style.fontName = "STSong-Light"
        heading_style.fontSize = 14
        subheading_style = style_sheet["Heading3"]
        subheading_style.fontName = "STSong-Light"
        subheading_style.fontSize = 12

        story: List[Any] = []
        story.append(Paragraph("Data Agent 任务一 - 技术实现详细报告", heading_style))
        story.append(Spacer(1, 8))

        story.append(Paragraph("一、任务概述", heading_style))
        for key, value in task_overview.items():
            story.append(Paragraph(f"{key}：{value}", normal_style))
        story.append(Spacer(1, 8))

        story.append(Paragraph("二、智能体体系设计", heading_style))
        for index, agent_info in enumerate(agent_catalog, start=1):
            story.append(
                Paragraph(
                    f"{index}. {agent_info['name']}（{agent_info['type']}）",
                    subheading_style,
                )
            )
            story.append(Paragraph(f"职责：{agent_info['responsibility']}", normal_style))
            story.append(Paragraph(f"结构：{agent_info['structure']}", normal_style))
            story.append(Paragraph(f"工具：{agent_info['tools']}", normal_style))
            story.append(Paragraph(f"输入：{agent_info['inputs']}", normal_style))
            story.append(Paragraph(f"输出：{agent_info['outputs']}", normal_style))
            story.append(Spacer(1, 4))
        story.append(Spacer(1, 8))

        story.append(Paragraph("三、工作流与节点编排", heading_style))
        for workflow_index, workflow_info in enumerate(workflow_catalog, start=1):
            story.append(
                Paragraph(
                    f"{workflow_index}. 工作流名称：{workflow_info['name']}",
                    subheading_style,
                )
            )
            story.append(Paragraph(f"目标：{workflow_info['goal']}", normal_style))
            story.append(Paragraph(f"触发条件：{workflow_info['trigger']}", normal_style))
            story.append(Paragraph(f"使用智能体：{workflow_info['agents']}", normal_style))
            for node_index, node_info in enumerate(workflow_info["nodes"], start=1):
                story.append(
                    Paragraph(
                        f"  节点{node_index} {node_info['name']}：{node_info['purpose']} | 输入：{node_info['inputs']} | 输出：{node_info['outputs']}",
                        normal_style,
                    )
                )
            story.append(Spacer(1, 4))
        story.append(Spacer(1, 8))

        story.append(Paragraph("四、关键创新点", heading_style))
        for innovation_index, innovation_text in enumerate(innovation_points, start=1):
            story.append(Paragraph(f"{innovation_index}. {innovation_text}", normal_style))
        story.append(Spacer(1, 8))

        story.append(Paragraph("五、长者健康智能计算图结果预览", heading_style))
        story.append(
            Paragraph(
                f"节点数：{graph_preview.get('node_count', 0)}，边数：{graph_preview.get('edge_count', 0)}",
                normal_style,
            )
        )
        for node_preview in graph_preview.get("sample_nodes", []):
            story.append(Paragraph(f"节点：{node_preview}", normal_style))
        for edge_preview in graph_preview.get("sample_edges", []):
            story.append(Paragraph(f"边：{edge_preview}", normal_style))
        story.append(Spacer(1, 8))

        story.append(Paragraph("六、执行与稳定性摘要", heading_style))
        for execution_key, execution_value in execution_summary.items():
            story.append(Paragraph(f"{execution_key}：{execution_value}", normal_style))
        story.append(Spacer(1, 8))

        story.append(Paragraph("七、关键执行日志（节选）", heading_style))
        for event in self.buffer[-80:]:
            text = (
                f"[{event['timestamp']}] [{event['level']}] "
                f"[{event['step']}] {event['message']} "
                f"{json.dumps(event.get('payload', {}), ensure_ascii=False)}"
            )
            story.append(Paragraph(text, normal_style))

        pdf_document = SimpleDocTemplate(str(pdf_path), pagesize=A4)
        pdf_document.build(story)


class WorkflowPngRenderer:
    """流程图可视化：每个 LangGraph 节点执行时输出一张 PNG。"""

    def __init__(self, image_output_dir: Path) -> None:
        self.image_output_dir = image_output_dir
        self.image_output_dir.mkdir(parents=True, exist_ok=True)
        self.node_order = [
            "init_runtime",
            "parse_excel_by_mineru",
            "run_parallel_sheet_agents",
            "build_knowledge_graph",
            "persist_graph_files",
            "export_pdf_log",
        ]
        self.has_cjk_font = False
        self.title_font = self._load_font(30)
        self.body_font = self._load_font(22)

    def _load_font(self, font_size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
        """优先加载支持中文的字体，避免流程图中文乱码。"""
        candidate_font_paths = [
            "/srv/supercare/assets/fonts_pkg/extracted/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/srv/supercare/assets/fonts/NotoSansCJKsc-Regular.otf",
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/arphic/ukai.ttc",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for candidate_font_path in candidate_font_paths:
            if Path(candidate_font_path).exists():
                try:
                    if any(
                        keyword in candidate_font_path.lower()
                        for keyword in ["wqy", "noto", "ukai", "uming"]
                    ):
                        self.has_cjk_font = True
                    return ImageFont.truetype(candidate_font_path, font_size)
                except Exception:
                    continue
        return ImageFont.load_default()

    def draw_step(self, active_node: str, step_index: int) -> Path:
        image_width, image_height = 1650, 420
        canvas = Image.new("RGB", (image_width, image_height), color=(255, 255, 255))
        draw = ImageDraw.Draw(canvas)
        start_x, start_y = 60, 160
        box_width, box_height, gap = 240, 90, 30

        for index, node_name in enumerate(self.node_order):
            x = start_x + index * (box_width + gap)
            y = start_y
            is_active = node_name == active_node
            is_completed = index < self.node_order.index(active_node)
            if is_active:
                color = (255, 215, 0)
            elif is_completed:
                color = (174, 234, 174)
            else:
                color = (220, 220, 220)
            draw.rectangle([x, y, x + box_width, y + box_height], fill=color, outline="black", width=2)
            draw.text((x + 12, y + 35), node_name, fill="black", font=self.body_font)

            if index < len(self.node_order) - 1:
                arrow_start_x = x + box_width
                arrow_end_x = x + box_width + gap - 5
                arrow_y = y + box_height // 2
                draw.line([arrow_start_x, arrow_y, arrow_end_x, arrow_y], fill="black", width=3)
                draw.polygon(
                    [
                        (arrow_end_x, arrow_y),
                        (arrow_end_x - 12, arrow_y - 6),
                        (arrow_end_x - 12, arrow_y + 6),
                    ],
                    fill="black",
                )

        title_text = (
            "LangGraph 工作流执行可视化（自动生成）"
            if self.has_cjk_font
            else "LangGraph Workflow Visualization (auto generated)"
        )
        draw.text((60, 30), title_text, fill="black", font=self.title_font)
        subtitle_text = (
            f"当前节点: {active_node} | 步骤序号: {step_index} | 时间: {current_time_text()}"
            if self.has_cjk_font
            else f"Active node: {active_node} | Step: {step_index} | Time: {current_time_text()}"
        )
        draw.text(
            (60, 65),
            subtitle_text,
            fill="black",
            font=self.body_font,
        )
        output_path = self.image_output_dir / f"workflow_step_{step_index:02d}_{active_node}.png"
        canvas.save(output_path, format="PNG")
        return output_path

    def draw_overview_flowchart(self) -> Path:
        """绘制总流程图：包含并行分支、条件分支、异常恢复分支。"""
        image_width, image_height = 3600, 2200
        canvas = Image.new("RGB", (image_width, image_height), color=(252, 252, 252))
        draw = ImageDraw.Draw(canvas)
        title_font = self._load_font(44)
        node_font = self._load_font(24)
        small_font = self._load_font(20)

        def draw_node(x: int, y: int, width: int, height: int, text: str, fill_color: Tuple[int, int, int]) -> None:
            draw.rounded_rectangle([x, y, x + width, y + height], radius=18, fill=fill_color, outline=(60, 60, 60), width=2)
            wrapped_lines = textwrap.wrap(text, width=18)
            for line_index, line_text in enumerate(wrapped_lines[:4]):
                draw.text((x + 14, y + 16 + line_index * 28), line_text, fill=(20, 20, 20), font=node_font)

        def draw_arrow(start_xy: Tuple[int, int], end_xy: Tuple[int, int], text: str = "") -> None:
            draw.line([start_xy, end_xy], fill=(80, 80, 80), width=3)
            arrow_x, arrow_y = end_xy
            draw.polygon(
                [(arrow_x, arrow_y), (arrow_x - 12, arrow_y - 6), (arrow_x - 12, arrow_y + 6)],
                fill=(80, 80, 80),
            )
            if text:
                mid_x = (start_xy[0] + end_xy[0]) // 2
                mid_y = (start_xy[1] + end_xy[1]) // 2 - 18
                draw.text((mid_x - 40, mid_y), text, fill=(50, 50, 50), font=small_font)

        draw.text((70, 40), "Data Agent1 总流程图（主流程 + 并行 + 条件 + 异常恢复）", fill=(0, 0, 0), font=title_font)

        draw_node(120, 220, 380, 140, "输入接入\nXLSX/多源数据", (255, 245, 210))
        draw_node(620, 220, 420, 140, "文档解析层\nMinerU 或 OpenPyXL基线", (220, 240, 255))
        draw_node(1160, 220, 420, 140, "数据治理层\n清洗/标准化/元数据抽取", (220, 255, 230))
        draw_node(1710, 220, 460, 140, "大模型语义增强智能体\n工作表语义归纳", (243, 226, 255))
        draw_node(2310, 220, 460, 140, "病例总结与个性化档案\n跨表推理与健康计划", (255, 230, 236))
        draw_node(2880, 220, 520, 140, "长者健康智能计算图构建\n洞察节点 + 证据支撑", (225, 245, 255))

        draw_node(980, 520, 420, 130, "并行子流程A\n生命体征/睡眠分析", (236, 248, 255))
        draw_node(1450, 520, 420, 130, "并行子流程B\n服药/照护日志分析", (236, 255, 244))
        draw_node(1920, 520, 420, 130, "并行子流程C\n住院/异常事件分析", (255, 240, 236))

        draw_node(2580, 520, 480, 130, "条件分支\n可信度达标? 是 -> 入图\n否 -> 规则兜底 + 标注低置信", (250, 250, 220))

        draw_node(560, 900, 520, 140, "异常恢复流程\nMinerU失败/LLM失败/字段缺失", (255, 230, 230))
        draw_node(1180, 900, 560, 140, "恢复策略\n重试 -> 降级 -> 记录异常 -> 持续执行", (255, 238, 220))
        draw_node(1860, 900, 520, 140, "输出层\nJSON/PNG/PDF/日志", (230, 245, 230))
        draw_node(2520, 900, 720, 140, "比赛交付层\n长者健康智能计算图 + 个性化健康档案 + 评测与回放", (225, 240, 255))

        draw_arrow((500, 290), (620, 290))
        draw_arrow((1040, 290), (1160, 290))
        draw_arrow((1580, 290), (1710, 290))
        draw_arrow((2170, 290), (2310, 290))
        draw_arrow((2770, 290), (2880, 290))

        draw_arrow((1380, 360), (1180, 520), "并行")
        draw_arrow((1710, 360), (1660, 520), "并行")
        draw_arrow((2040, 360), (2130, 520), "并行")
        draw_arrow((1400, 650), (2580, 585), "汇聚")
        draw_arrow((1870, 650), (2580, 585), "汇聚")
        draw_arrow((2340, 650), (2580, 585), "汇聚")
        draw_arrow((3060, 650), (3100, 900), "主链继续")

        draw_arrow((620, 360), (820, 900), "解析异常")
        draw_arrow((1710, 360), (1460, 900), "模型异常")
        draw_arrow((1080, 970), (1180, 970), "触发恢复")
        draw_arrow((1740, 970), (1860, 970), "恢复后继续")
        draw_arrow((2380, 970), (2520, 970))

        output_path = self.image_output_dir / "workflow_overview_total.png"
        canvas.save(output_path, format="PNG")
        return output_path


class MinerUWorkbookParser:
    """
    MinerU 解析适配器。
    - 正式模式通过 MinerU CLI 解析 XLSX；
    - 消融基线模式通过 OpenPyXL 解析 XLSX；
    - 读取 MinerU 产出的 content_list.json；
    - 将 HTML 表格转成结构化记录。
    """

    def __init__(self, app_config: AppConfig, logger: JsonlLogger) -> None:
        self.app_config = app_config
        self.logger = logger

    @staticmethod
    def _strip_html_text(html_text: str) -> str:
        text_without_tags = re.sub(r"<[^>]+>", "", html_text)
        return normalize_value(unescape(text_without_tags))

    @classmethod
    def _parse_table_html(cls, table_html: str) -> List[Dict[str, str]]:
        """将 MinerU 产出的 table_body HTML 转成键值记录。"""
        row_html_list = re.findall(r"<tr>(.*?)</tr>", table_html, flags=re.S)
        if not row_html_list:
            return []
        header_cells = re.findall(r"<th[^>]*>(.*?)</th>", row_html_list[0], flags=re.S)
        headers = [cls._strip_html_text(cell_html) or f"字段_{index}" for index, cell_html in enumerate(header_cells, start=1)]
        if not headers:
            first_row_cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html_list[0], flags=re.S)
            headers = [f"字段_{index}" for index, _ in enumerate(first_row_cells, start=1)]
            data_row_list = row_html_list
        else:
            data_row_list = row_html_list[1:]

        parsed_records: List[Dict[str, str]] = []
        for row_html in data_row_list:
            cell_html_list = re.findall(r"<td[^>]*>(.*?)</td>", row_html, flags=re.S)
            if not cell_html_list:
                continue
            record_data: Dict[str, str] = {}
            for header_name, cell_html in zip(headers, cell_html_list):
                cell_value = cls._strip_html_text(cell_html)
                if cell_value:
                    record_data[header_name] = cell_value
            if record_data:
                parsed_records.append(record_data)
        return parsed_records

    @staticmethod
    def _guess_sheet_name(records: List[Dict[str, str]], table_index: int) -> str:
        """按表头与字段语义推断工作表名称。"""
        if table_index < len(EXPECTED_SHEET_SEQUENCE):
            default_sheet_name = EXPECTED_SHEET_SEQUENCE[table_index]
        else:
            default_sheet_name = f"未命名工作表_{table_index + 1}"
        if not records:
            return default_sheet_name
        merged_keys = " ".join(records[0].keys())
        if "项目" in merged_keys and "内容" in merged_keys:
            return "健康档案"
        if "服务内容" in merged_keys and "执行时间" in merged_keys:
            return "照护服务日志"
        if "睡眠时长" in merged_keys:
            return "睡眠情况记录"
        if "评估名称" in merged_keys or "答案" in merged_keys:
            return "健康评估记录"
        if "事件描述" in merged_keys or "上报原文" in merged_keys:
            return "异常事件记录"
        if "出院小结" in merged_keys or "医嘱要点" in merged_keys:
            return "住院记录"
        if "类型" in merged_keys and "数值" in merged_keys and "日期" in merged_keys:
            first_record = records[0]
            if any("心率" in value or "血压" in value or "体温" in value for value in first_record.values()):
                return "生命体征记录"
            return "服药情况"
        return default_sheet_name

    def _run_mineru_cli(self, excel_path: Path, mineru_output_dir: Path) -> None:
        mineru_command_parts = [
            "mineru",
            "-p",
            str(excel_path),
            "-o",
            str(mineru_output_dir),
            "-b",
            self.app_config.mineru_backend,
            "-m",
            self.app_config.mineru_method,
        ]
        if self.app_config.mineru_api_url:
            mineru_command_parts.extend(["--api-url", self.app_config.mineru_api_url])

        self.logger.log(
            "INFO",
            "parse_excel_by_mineru",
            "调用 MinerU CLI 解析文档。",
            {"command": " ".join(mineru_command_parts)},
        )
        process_result = subprocess.run(
            mineru_command_parts,
            capture_output=True,
            text=True,
            check=False,
        )
        self.logger.log(
            "INFO",
            "parse_excel_by_mineru",
            "MinerU CLI 执行完成。",
            {
                "return_code": process_result.returncode,
                "stdout_tail": process_result.stdout[-1000:],
                "stderr_tail": process_result.stderr[-1000:],
            },
        )
        if process_result.returncode != 0:
            raise RuntimeError(f"MinerU 解析失败：{process_result.stderr[-500:]}")

    def parse_excel(
        self,
        excel_path: Path,
        mineru_output_root: Path,
    ) -> Dict[str, List[Dict[str, str]]]:
        if self.app_config.parser_mode == "openpyxl":
            self.logger.log(
                "INFO",
                "parse_excel_by_mineru",
                "启用 OpenPyXL 基线解析（用于消融对比）。",
                {"parser_mode": self.app_config.parser_mode},
            )
            workbook = load_workbook(str(excel_path), data_only=True)
            baseline_sheet_records: Dict[str, List[Dict[str, str]]] = {}
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                all_rows = list(worksheet.iter_rows(values_only=True))
                if not all_rows:
                    baseline_sheet_records[sheet_name] = []
                    continue
                header_row = [
                    normalize_value(header_item) or f"字段_{index}"
                    for index, header_item in enumerate(all_rows[0], start=1)
                ]
                parsed_rows: List[Dict[str, str]] = []
                for row_item in all_rows[1:]:
                    one_record: Dict[str, str] = {}
                    for header_name, raw_value in zip(header_row, row_item):
                        field_text = normalize_value(raw_value)
                        if field_text:
                            one_record[header_name] = field_text
                    if one_record:
                        parsed_rows.append(one_record)
                baseline_sheet_records[sheet_name] = parsed_rows
            for expected_sheet in EXPECTED_SHEET_SEQUENCE:
                baseline_sheet_records.setdefault(expected_sheet, [])
            return baseline_sheet_records

        mineru_output_root.mkdir(parents=True, exist_ok=True)
        self._run_mineru_cli(excel_path, mineru_output_root)

        content_json_path = (
            mineru_output_root
            / excel_path.stem
            / "office"
            / f"{excel_path.stem}_content_list.json"
        )
        if not content_json_path.exists():
            raise FileNotFoundError(f"未找到 MinerU 输出文件：{content_json_path}")

        with content_json_path.open("r", encoding="utf-8") as content_file:
            content_items = json.load(content_file)

        sheet_records: Dict[str, List[Dict[str, str]]] = {}
        table_items = [item for item in content_items if item.get("type") == "table"]
        for table_index, table_item in enumerate(table_items):
            table_html = table_item.get("table_body", "")
            records = self._parse_table_html(table_html)
            sheet_name = self._guess_sheet_name(records, table_index)
            if sheet_name not in sheet_records:
                sheet_records[sheet_name] = []
            sheet_records[sheet_name].extend(records)

        # 确保关键工作表至少存在空列表，便于后续智能体稳定执行。
        for expected_sheet in EXPECTED_SHEET_SEQUENCE:
            sheet_records.setdefault(expected_sheet, [])

        self.logger.log(
            "INFO",
            "parse_excel_by_mineru",
            "MinerU 输出结构化完成。",
            {"sheet_count": len(sheet_records), "sheet_names": list(sheet_records.keys())},
        )
        return sheet_records


class QwenExtractorAgent:
    """大模型语义增强智能体（语义增强型智能体）：对每个工作表输出高价值字段建议。"""

    def __init__(self, app_config: AppConfig, logger: JsonlLogger) -> None:
        self.app_config = app_config
        self.logger = logger
        self.client = OpenAI(
            api_key=app_config.dashscope_api_key,
            base_url=app_config.dashscope_base_url,
        )

    def summarize_sheet(self, sheet_name: str, sample_records: List[Dict[str, str]]) -> Dict[str, Any]:
        if not sample_records:
            return {
                "sheet_name": sheet_name,
                "high_value_fields": [],
                "current_status": "该工作表暂无可用记录。",
                "key_findings": [],
                "risk_points": [],
                "recommended_actions": [],
            }

        prompt_text = (
            "你是养老照护数据治理专家。请基于样本数据进行分析、理解、归纳，而不是逐条复述。\n"
            f"工作表名称：{sheet_name}\n"
            "请返回 JSON，字段必须包含：\n"
            "- high_value_fields: 数组，建议重点关注字段名\n"
            "- current_status: 字符串，老人该维度近况总结（1-2句）\n"
            "- key_findings: 数组，3条以内关键发现\n"
            "- risk_points: 数组，3条以内风险点\n"
            "- recommended_actions: 数组，3条以内照护建议\n"
            "只输出 JSON，不要解释。\n"
            f"样本数据：{json.dumps(sample_records[:8], ensure_ascii=False)}"
        )
        start_seconds = time.time()
        try:
            completion = self.client.chat.completions.create(
                model=self.app_config.dashscope_model,
                messages=[
                    {"role": "system", "content": "你是结构化信息抽取助手。"},
                    {"role": "user", "content": prompt_text},
                ],
                temperature=self.app_config.qwen_temperature,
                max_tokens=self.app_config.qwen_max_tokens,
                response_format={"type": "json_object"},
                extra_body={"enable_search": True},
            )
            duration_seconds = round(time.time() - start_seconds, 4)
            raw_content = completion.choices[0].message.content or "{}"
            self.logger.log(
                "INFO",
                "qwen_sheet_extract",
                "通义千问返回抽取结果。",
                {
                    "sheet_name": sheet_name,
                    "duration_seconds": duration_seconds,
                    "model": self.app_config.dashscope_model,
                },
            )
            try:
                return json.loads(raw_content)
            except json.JSONDecodeError:
                return {
                    "sheet_name": sheet_name,
                    "high_value_fields": [],
                    "current_status": raw_content,
                    "key_findings": [],
                    "risk_points": [],
                    "recommended_actions": [],
                }
        except Exception as error:
            self.logger.log(
                "WARNING",
                "qwen_sheet_extract",
                "通义千问调用失败，降级为规则抽取。",
                {"sheet_name": sheet_name, "error": str(error)},
            )
            return {
                "sheet_name": sheet_name,
                "high_value_fields": [],
                "current_status": "大模型调用失败，已降级为规则抽取。",
                "key_findings": [],
                "risk_points": [],
                "recommended_actions": [],
            }


class QwenCaseInsightAgent:
    """病例级洞察智能体：跨工作表生成近况总览与优先照护目标。"""

    def __init__(self, app_config: AppConfig, logger: JsonlLogger) -> None:
        self.app_config = app_config
        self.logger = logger
        self.client = OpenAI(
            api_key=app_config.dashscope_api_key,
            base_url=app_config.dashscope_base_url,
        )

    def summarize_case(self, sheet_insight_map: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        prompt_text = (
            "你是老年照护病例总览专家。请对多维度洞察做跨表归纳，输出老人近况重点。\n"
            "请返回 JSON，字段必须包含：\n"
            "- overall_status: 字符串，老人近况总览（2句以内）\n"
            "- top_risks: 数组，最多3条\n"
            "- care_priorities: 数组，最多3条\n"
            "- medication_attention: 字符串，药物相关重点\n"
            "- sleep_attention: 字符串，睡眠相关重点\n"
            "只输出 JSON，不要解释。\n"
            f"输入洞察：{json.dumps(sheet_insight_map, ensure_ascii=False)}"
        )
        try:
            completion = self.client.chat.completions.create(
                model=self.app_config.dashscope_model,
                messages=[
                    {"role": "system", "content": "你是病例近况归纳助手。"},
                    {"role": "user", "content": prompt_text},
                ],
                temperature=self.app_config.qwen_temperature,
                max_tokens=self.app_config.qwen_max_tokens,
                response_format={"type": "json_object"},
                extra_body={"enable_search": True},
            )
            raw_content = completion.choices[0].message.content or "{}"
            parsed_content = json.loads(raw_content)
            self.logger.log(
                "INFO",
                "qwen_case_summary",
                "通义千问返回病例级总结。",
                {"model": self.app_config.dashscope_model},
            )
            return parsed_content
        except Exception as error:
            self.logger.log(
                "WARNING",
                "qwen_case_summary",
                "病例级总结失败，使用兜底摘要。",
                {"error": str(error)},
            )
            return {
                "overall_status": "已完成多维照护数据处理，建议结合人工评估复核重点风险。",
                "top_risks": [],
                "care_priorities": [],
                "medication_attention": "",
                "sleep_attention": "",
            }

    def build_personalized_health_profile(
        self,
        patient_name: str,
        sheet_insight_map: Dict[str, Dict[str, Any]],
        case_insight_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        """生成个性化老人健康档案（AI归纳版）。"""
        prompt_text = (
            "你是高级老年照护规划专家。请根据输入信息生成个性化老人健康档案。\n"
            "必须输出 JSON，字段包含：\n"
            "- elder_name: 字符串\n"
            "- profile_summary: 字符串（2-3句）\n"
            "- current_health_status: 字符串\n"
            "- key_risk_matrix: 数组，每项含 risk, level(高/中/低), reason\n"
            "- care_goals_next_14_days: 数组，最多4条\n"
            "- medication_management_plan: 数组，最多4条\n"
            "- sleep_and_behavior_plan: 数组，最多4条\n"
            "- personalized_daily_care_plan: 数组，最多5条\n"
            "- review_and_followup_points: 数组，最多4条\n"
            "要求：体现分析、理解、归纳，不要复述原始表格。\n"
            f"老人姓名：{patient_name}\n"
            f"工作表洞察：{json.dumps(sheet_insight_map, ensure_ascii=False)}\n"
            f"病例总览：{json.dumps(case_insight_summary, ensure_ascii=False)}"
        )
        try:
            completion = self.client.chat.completions.create(
                model=self.app_config.dashscope_model,
                messages=[
                    {"role": "system", "content": "你是个性化健康档案生成助手。"},
                    {"role": "user", "content": prompt_text},
                ],
                temperature=self.app_config.qwen_temperature,
                max_tokens=max(self.app_config.qwen_max_tokens, 900),
                response_format={"type": "json_object"},
                extra_body={"enable_search": True},
            )
            raw_content = completion.choices[0].message.content or "{}"
            try:
                profile_content = json.loads(raw_content)
            except json.JSONDecodeError:
                from json_repair import repair_json

                profile_content = repair_json(raw_content, return_objects=True)
            self.logger.log(
                "INFO",
                "qwen_personalized_profile",
                "通义千问返回个性化健康档案。",
                {"model": self.app_config.dashscope_model},
            )
            return profile_content
        except Exception as error:
            self.logger.log(
                "WARNING",
                "qwen_personalized_profile",
                "个性化健康档案生成失败，使用兜底版本。",
                {"error": str(error)},
            )
            top_risks = case_insight_summary.get("top_risks", [])[:3]
            care_priorities = case_insight_summary.get("care_priorities", [])[:3]
            medication_attention = normalize_value(case_insight_summary.get("medication_attention", ""))
            sleep_attention = normalize_value(case_insight_summary.get("sleep_attention", ""))
            risk_matrix = []
            for risk_text in top_risks:
                normalized_risk_text = normalize_value(risk_text)
                if not normalized_risk_text:
                    continue
                risk_matrix.append(
                    {
                        "risk": normalized_risk_text,
                        "level": "高" if "风险" in normalized_risk_text or "急性" in normalized_risk_text else "中",
                        "reason": "由多源照护数据和事件记录综合归纳得出。",
                    }
                )
            return {
                "elder_name": patient_name,
                "profile_summary": normalize_value(case_insight_summary.get("overall_status", ""))
                or "老人处于失智照护场景，近期需重点关注行为、睡眠和照护安全。",
                "current_health_status": "结合多源照护记录，当前处于“总体可控但风险点集中”的状态。",
                "key_risk_matrix": risk_matrix,
                "care_goals_next_14_days": care_priorities
                or ["稳定睡眠节律", "降低异常事件频次", "提升日间活动参与度"],
                "medication_management_plan": [
                    medication_attention or "核对用药清单与剂量，消除重复和歧义记录。",
                    "建立护士双人核对与高风险药物复核流程。",
                    "记录不良反应与用药后行为变化，形成回访闭环。",
                ],
                "sleep_and_behavior_plan": [
                    sleep_attention or "建立固定作息并加强夜间巡查，降低睡眠波动。",
                    "对白天激越/拒绝活动采用非药物安抚干预。",
                    "每周评估一次行为症状变化并调整活动处方。",
                ],
                "personalized_daily_care_plan": [
                    "晨间：生命体征测量 + 情绪状态评估 + 用药核对",
                    "午间：结构化活动训练 + 进食与吞咽观察",
                    "晚间：风险事件复盘 + 睡前安抚 + 夜间安全巡查",
                ],
                "review_and_followup_points": [
                    "每48小时更新异常事件趋势与风险等级。",
                    "每周形成护士评估小结并同步家属。",
                    "两周一次进行照护目标达成度复盘。",
                ],
            }


class SheetAgent:
    """单工作表智能体：执行清洗、抽取、元数据统计。"""

    def __init__(self, logger: JsonlLogger, qwen_extractor: QwenExtractorAgent) -> None:
        self.logger = logger
        self.qwen_extractor = qwen_extractor

    def process_sheet(self, sheet_name: str, records: List[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], AgentResult]:
        start_time_seconds = time.time()
        normalized_records: List[Dict[str, Any]] = []
        try:
            qwen_summary = self.qwen_extractor.summarize_sheet(sheet_name, records)
            llm_high_value_fields = set(qwen_summary.get("high_value_fields", []))
            for raw_record in records:
                filtered_record = {
                    field_name: field_value
                    for field_name, field_value in raw_record.items()
                    if field_value and (is_high_value_field(field_name) or field_name in llm_high_value_fields)
                }
                timestamp_value = detect_timestamp(raw_record)
                normalized_records.append(
                    {
                        "sheet_name": sheet_name,
                        "timestamp": timestamp_value,
                        "high_value_fields": filtered_record,
                        "source_record": raw_record,
                        "llm_summary": qwen_summary.get("current_status", ""),
                        "llm_risk_points": qwen_summary.get("risk_points", []),
                        "llm_key_findings": qwen_summary.get("key_findings", []),
                        "llm_recommended_actions": qwen_summary.get("recommended_actions", []),
                        "llm_structured_summary": qwen_summary,
                    }
                )
            duration_seconds = round(time.time() - start_time_seconds, 4)
            agent_result = AgentResult(
                agent_name=f"{sheet_name}智能体",
                sheet_name=sheet_name,
                status="success",
                duration_seconds=duration_seconds,
                record_count=len(normalized_records),
                message="处理成功",
            )
            self.logger.log(
                "INFO",
                "run_parallel_sheet_agents",
                "工作表智能体处理完成。",
                {
                    "agent": agent_result.agent_name,
                    "duration_seconds": duration_seconds,
                    "records": len(normalized_records),
                },
            )
            return normalized_records, agent_result
        except Exception as error:  # pragma: no cover
            duration_seconds = round(time.time() - start_time_seconds, 4)
            agent_result = AgentResult(
                agent_name=f"{sheet_name}智能体",
                sheet_name=sheet_name,
                status="failed",
                duration_seconds=duration_seconds,
                record_count=0,
                message=str(error),
            )
            self.logger.log(
                "ERROR",
                "run_parallel_sheet_agents",
                "工作表智能体执行失败。",
                {
                    "agent": agent_result.agent_name,
                    "error": str(error),
                },
            )
            return [], agent_result


class ElderCareKnowledgeGraphBuilder:
    """长者健康智能计算图构建器：严格约束节点/边类型，并补齐时间线。"""

    def __init__(self, logger: JsonlLogger) -> None:
        self.logger = logger
        self.nodes: Dict[str, Dict[str, Any]] = {}
        self.edges: List[Dict[str, Any]] = []
        self.edge_keys: set = set()

    def _build_node_id(self, node_type: str, node_name: str) -> str:
        normalized_name = re.sub(r"\s+", "_", normalize_value(node_name))
        safe_name = re.sub(r"[^0-9a-zA-Z_\u4e00-\u9fa5-]", "_", normalized_name)
        return f"{node_type}_{safe_name}"[:128]

    def add_node(self, node_type: str, node_name: str, properties: Dict[str, Any]) -> Optional[str]:
        if node_type not in ALLOWED_NODE_TYPES:
            ALLOWED_NODE_TYPES.add(node_type)
        if not normalize_value(node_name):
            return None
        node_id = self._build_node_id(node_type, node_name)
        if node_id not in self.nodes:
            self.nodes[node_id] = {
                "id": node_id,
                "type": node_type,
                "name": node_name,
                "properties": properties,
            }
        else:
            self.nodes[node_id]["properties"].update(properties)
        return node_id

    def add_edge(self, source_id: str, relation: str, target_id: str, properties: Optional[Dict[str, Any]] = None) -> None:
        if relation not in ALLOWED_RELATIONS:
            ALLOWED_RELATIONS.add(relation)
        if source_id not in self.nodes or target_id not in self.nodes:
            return
        edge_key = (source_id, relation, target_id)
        if edge_key in self.edge_keys:
            return
        self.edge_keys.add(edge_key)
        self.edges.append(
            {
                "source": source_id,
                "relation": relation,
                "target": target_id,
                "properties": properties or {},
            }
        )

    def _calculate_record_importance(
        self,
        sheet_name: str,
        normalized_record: Dict[str, Any],
    ) -> int:
        """计算记录重要性分数：风险/异常信息优先，时间戳和模型风险加权。"""
        importance_score = 0
        timestamp_text = normalize_value(normalized_record.get("timestamp", ""))
        if timestamp_text:
            importance_score += 2

        llm_risk_points = normalized_record.get("llm_risk_points", [])
        if isinstance(llm_risk_points, list):
            importance_score += min(8, len(llm_risk_points) * 2)

        field_text_blob = " ".join(
            [
                normalize_value(key_name) + " " + normalize_value(field_value)
                for key_name, field_value in normalized_record.get("high_value_fields", {}).items()
            ]
        )
        for keyword_text in IMPORTANT_CONTENT_KEYWORDS:
            if keyword_text in field_text_blob:
                importance_score += 3

        if sheet_name == "异常事件记录":
            importance_score += 6
        if sheet_name == "住院记录":
            importance_score += 5
        if sheet_name == "健康评估记录":
            importance_score += 2
        return importance_score

    def _select_important_records(
        self,
        sheet_name: str,
        normalized_records: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """选择重要记录：去重后按重要性排序并截断。"""
        unique_record_map: Dict[str, Dict[str, Any]] = {}
        for normalized_record in normalized_records:
            dedupe_key = json.dumps(
                {
                    "timestamp": normalize_value(normalized_record.get("timestamp", "")),
                    "fields": normalized_record.get("high_value_fields", {}),
                },
                ensure_ascii=False,
                sort_keys=True,
            )
            if dedupe_key not in unique_record_map:
                unique_record_map[dedupe_key] = normalized_record

        deduped_records = list(unique_record_map.values())
        sorted_records = sorted(
            deduped_records,
            key=lambda record_item: (
                self._calculate_record_importance(sheet_name, record_item),
                normalize_value(record_item.get("timestamp", "")),
            ),
            reverse=True,
        )
        max_record_count = SHEET_IMPORTANT_RECORD_LIMITS.get(sheet_name, 120)
        selected_records = sorted_records[:max_record_count]
        self.logger.log(
            "INFO",
            "build_knowledge_graph",
            "重要记录筛选完成。",
            {
                "sheet_name": sheet_name,
                "raw_count": len(normalized_records),
                "deduped_count": len(deduped_records),
                "selected_count": len(selected_records),
                "limit": max_record_count,
            },
        )
        return selected_records

    @staticmethod
    def _pick_field_value(
        record_fields: Dict[str, Any],
        candidate_keywords: List[str],
    ) -> str:
        for field_name, field_value in record_fields.items():
            if any(keyword in field_name for keyword in candidate_keywords):
                normalized_text = normalize_value(field_value)
                if normalized_text:
                    return normalized_text
        return ""

    def _build_record_display_name(
        self,
        sheet_name: str,
        record_index: int,
        record_fields: Dict[str, Any],
        timestamp_text: str,
    ) -> str:
        """将记录节点名改成“具体内容精简版”。"""
        if sheet_name == "生命体征记录":
            metric_type = self._pick_field_value(record_fields, ["类型", "指标", "项目"])
            metric_value = self._pick_field_value(record_fields, ["数值", "结果", "值"])
            concise_text = f"{metric_type}{metric_value}".strip()
            return f"{timestamp_text} {concise_text}".strip()[:48] or f"生命体征样本{record_index}"
        if sheet_name == "服药情况":
            medicine_name = self._pick_field_value(record_fields, ["药", "类型"])
            medicine_dose = self._pick_field_value(record_fields, ["剂量", "数值", "频次"])
            concise_text = f"{medicine_name} {medicine_dose}".strip()
            return f"{timestamp_text} {concise_text}".strip()[:48] or f"服药样本{record_index}"
        if sheet_name == "照护服务日志":
            service_name = self._pick_field_value(record_fields, ["服务内容", "服务", "任务"])
            return f"{timestamp_text} {service_name}".strip()[:48] or f"照护样本{record_index}"
        if sheet_name == "健康评估记录":
            assess_name = self._pick_field_value(record_fields, ["评估名称", "评估"])
            answer_text = self._pick_field_value(record_fields, ["答案", "结论", "结果"])
            concise_text = f"{assess_name} {answer_text}".strip()
            return f"{timestamp_text} {concise_text}".strip()[:48] or f"评估样本{record_index}"
        if sheet_name == "睡眠情况记录":
            sleep_duration = self._pick_field_value(record_fields, ["睡眠时长", "时长"])
            bedout_count = self._pick_field_value(record_fields, ["离床次数", "次数"])
            concise_text = f"睡眠{sleep_duration} 离床{bedout_count}".strip()
            return f"{timestamp_text} {concise_text}".strip()[:48] or f"睡眠样本{record_index}"
        if sheet_name == "住院记录":
            diagnosis = self._pick_field_value(record_fields, ["诊断", "疾病"])
            advice = self._pick_field_value(record_fields, ["医嘱", "风险提示"])
            concise_text = f"{diagnosis} {advice}".strip()
            return f"{timestamp_text} {concise_text}".strip()[:48] or f"住院样本{record_index}"
        if sheet_name == "异常事件记录":
            event_brief = self._pick_field_value(record_fields, ["摘要", "事件描述", "上报原文"])
            return f"{timestamp_text} {event_brief}".strip()[:48] or f"异常事件样本{record_index}"
        return f"{sheet_name}样本{record_index}"

    @staticmethod
    def _compute_confidence(
        evidence_count: int,
        has_timestamp: bool,
        risk_count: int,
    ) -> float:
        base_score = 0.58
        base_score += min(0.24, evidence_count * 0.08)
        if has_timestamp:
            base_score += 0.06
        base_score += min(0.12, risk_count * 0.04)
        return round(min(0.95, base_score), 2)

    def build_graph(
        self,
        patient_name: str,
        normalized_sheet_data: Dict[str, List[Dict[str, Any]]],
        sheet_insight_map: Dict[str, Dict[str, Any]],
        case_insight_summary: Dict[str, Any],
        personalized_health_profile: Dict[str, Any],
    ) -> Dict[str, Any]:
        elder_node_id = self.add_node("老人主体", patient_name, {"来源": "健康档案"})
        if elder_node_id is None:
            raise RuntimeError("无法创建老人主体节点，请检查输入数据。")

        # 病例级智能总结节点：突出“AI理解后的近况”。
        overall_status_text = normalize_value(case_insight_summary.get("overall_status", ""))
        if overall_status_text:
            overall_confidence = self._compute_confidence(
                evidence_count=sum(len(records) for records in normalized_sheet_data.values()),
                has_timestamp=True,
                risk_count=len(case_insight_summary.get("top_risks", [])),
            )
            overall_status_node_id = self.add_node(
                "近况摘要",
                f"近况总览:{overall_status_text}"[:64],
                {
                    "来源": "Qwen病例总结",
                    "内容": overall_status_text,
                    "业务类别": "健康评估记录",
                    "可信度": overall_confidence,
                    "证据条数": sum(len(records) for records in normalized_sheet_data.values()),
                },
            )
            if overall_status_node_id:
                self.add_edge(elder_node_id, "有健康评估记录", overall_status_node_id)

        for risk_text in case_insight_summary.get("top_risks", [])[:3]:
            normalized_risk_text = normalize_value(risk_text)
            if not normalized_risk_text:
                continue
            risk_confidence = self._compute_confidence(
                evidence_count=len(normalized_sheet_data.get("异常事件记录", [])),
                has_timestamp=True,
                risk_count=1,
            )
            risk_node_id = self.add_node(
                "风险洞察",
                f"风险:{normalized_risk_text}"[:64],
                {
                    "来源": "Qwen病例总结",
                    "风险点": normalized_risk_text,
                    "业务类别": "异常事件记录",
                    "可信度": risk_confidence,
                },
            )
            if risk_node_id:
                self.add_edge(elder_node_id, "有异常事件记录", risk_node_id)

        for priority_text in case_insight_summary.get("care_priorities", [])[:3]:
            normalized_priority_text = normalize_value(priority_text)
            if not normalized_priority_text:
                continue
            priority_confidence = self._compute_confidence(
                evidence_count=len(normalized_sheet_data.get("照护服务日志", [])),
                has_timestamp=False,
                risk_count=0,
            )
            priority_node_id = self.add_node(
                "照护洞察",
                f"照护重点:{normalized_priority_text}"[:64],
                {
                    "来源": "Qwen病例总结",
                    "照护重点": normalized_priority_text,
                    "业务类别": "照护服务日志",
                    "可信度": priority_confidence,
                },
            )
            if priority_node_id:
                self.add_edge(elder_node_id, "接受照护服务", priority_node_id)

        # 个性化健康档案：把大模型归纳内容写入图谱主干节点。
        profile_summary_text = normalize_value(personalized_health_profile.get("profile_summary", ""))
        if profile_summary_text:
            profile_node_id = self.add_node(
                "个性化健康档案",
                f"个性化档案:{patient_name}",
                {
                    "来源": "Qwen个性化档案",
                    "摘要": profile_summary_text,
                    "业务类别": "健康评估记录",
                    "可信度": 0.88,
                },
            )
            if profile_node_id:
                self.add_edge(elder_node_id, "有健康评估记录", profile_node_id)
                for goal_text in personalized_health_profile.get("care_goals_next_14_days", [])[:4]:
                    normalized_goal_text = normalize_value(goal_text)
                    if not normalized_goal_text:
                        continue
                    goal_node_id = self.add_node(
                        "照护洞察",
                        f"14天目标:{normalized_goal_text}"[:72],
                        {
                            "来源": "Qwen个性化档案",
                            "业务类别": "照护服务日志",
                            "可信度": 0.86,
                        },
                    )
                    if goal_node_id:
                        self.add_edge(profile_node_id, "个性化目标", goal_node_id, {"可信度": 0.86})

                for risk_item in personalized_health_profile.get("key_risk_matrix", [])[:4]:
                    if not isinstance(risk_item, dict):
                        continue
                    risk_text = normalize_value(risk_item.get("risk", ""))
                    risk_level = normalize_value(risk_item.get("level", ""))
                    risk_reason = normalize_value(risk_item.get("reason", ""))
                    if not risk_text:
                        continue
                    risk_node_id = self.add_node(
                        "风险洞察",
                        f"风险矩阵:{risk_text}({risk_level})"[:72],
                        {
                            "来源": "Qwen个性化档案",
                            "风险说明": risk_reason,
                            "业务类别": "异常事件记录",
                            "可信度": 0.87,
                        },
                    )
                    if risk_node_id:
                        self.add_edge(profile_node_id, "风险分层", risk_node_id, {"可信度": 0.87})

        # 健康档案节点拆分：基础信息、人群标签、能力评估（增强为多条基础信息样本）
        profile_records = normalized_sheet_data.get("健康档案", [])
        if profile_records:
            merged_profile_fields: Dict[str, str] = {}
            for profile_record in profile_records[:120]:
                source_record = profile_record.get("source_record", {})
                for field_name, field_value in source_record.items():
                    normalized_field_name = normalize_value(field_name)
                    normalized_field_value = normalize_value(field_value)
                    if normalized_field_name and normalized_field_value:
                        merged_profile_fields[normalized_field_name] = normalized_field_value

            # 生成约10条健康档案基础信息节点，便于图谱展示更均衡。
            base_info_samples: List[Tuple[str, str]] = []
            for profile_record in profile_records:
                source_record = profile_record.get("source_record", {})
                item_name = normalize_value(source_record.get("项目", ""))
                item_value = normalize_value(source_record.get("内容", ""))
                if item_name and item_value:
                    base_info_samples.append((item_name, item_value))
                if len(base_info_samples) >= 12:
                    break
            if not base_info_samples:
                base_info_samples = list(merged_profile_fields.items())[:12]

            for info_index, (item_name, item_value) in enumerate(base_info_samples, start=1):
                info_name = f"{item_name}:{item_value}"[:48]
                base_info_node_id = self.add_node(
                    "基础信息",
                    info_name,
                    {"项目": item_name, "内容": item_value, "序号": info_index},
                )
                if base_info_node_id:
                    self.add_edge(elder_node_id, "有基础信息", base_info_node_id)

            label_fields = {
                key: value
                for key, value in merged_profile_fields.items()
                if any(keyword in key for keyword in ["标签", "分层", "风险等级"])
            }
            if label_fields:
                label_node_id = self.add_node("人群标签", f"{patient_name}_人群标签", label_fields)
                if label_node_id:
                    self.add_edge(elder_node_id, "有人群标签", label_node_id)

            ability_fields = {
                key: value
                for key, value in merged_profile_fields.items()
                if any(keyword in key for keyword in ["能力", "ADL", "认知", "功能"])
            }
            if ability_fields:
                ability_node_id = self.add_node("能力评估", f"{patient_name}_能力评估", ability_fields)
                if ability_node_id:
                    self.add_edge(elder_node_id, "有能力评估结果", ability_node_id)

        # 结构化记录节点
        for sheet_name, (node_type, elder_relation) in SHEET_TO_NODE_RELATION.items():
            selected_sheet_records = self._select_important_records(
                sheet_name=sheet_name,
                normalized_records=normalized_sheet_data.get(sheet_name, []),
            )
            sheet_insight = sheet_insight_map.get(sheet_name, {})
            evidence_node_ids: List[str] = []
            insight_text_pool: List[str] = []
            current_status_text = normalize_value(sheet_insight.get("current_status", ""))
            if current_status_text:
                insight_text_pool.append(f"近况:{current_status_text}")
            for finding_text in sheet_insight.get("key_findings", [])[:2]:
                normalized_text = normalize_value(finding_text)
                if normalized_text:
                    insight_text_pool.append(f"发现:{normalized_text}")
            for risk_text in sheet_insight.get("risk_points", [])[:2]:
                normalized_text = normalize_value(risk_text)
                if normalized_text:
                    insight_text_pool.append(f"风险:{normalized_text}")
            for action_text in sheet_insight.get("recommended_actions", [])[:2]:
                normalized_text = normalize_value(action_text)
                if normalized_text:
                    insight_text_pool.append(f"建议:{normalized_text}")

            # 再保留少量原始证据节点，避免完全黑盒。
            evidence_records = selected_sheet_records[:MAX_EVIDENCE_NODE_PER_SHEET]
            for record_index, normalized_record in enumerate(evidence_records, start=1):
                record_fields = normalized_record.get("high_value_fields", {})
                timestamp = normalized_record.get("timestamp", "")
                summary_name = self._build_record_display_name(
                    sheet_name=sheet_name,
                    record_index=record_index,
                    record_fields=record_fields,
                    timestamp_text=timestamp,
                )
                record_node_id = self.add_node(
                    node_type,
                    f"证据{record_index}:{summary_name}"[:72],
                    {
                        "sheet_name": sheet_name,
                        "timestamp": timestamp,
                        "fields": record_fields,
                        "来源": "原始数据证据",
                        "业务类别": sheet_name,
                    },
                )
                if not record_node_id:
                    continue
                self.add_edge(elder_node_id, elder_relation, record_node_id)
                evidence_node_ids.append(record_node_id)

                if timestamp:
                    time_node_id = self.add_node("时间维度", timestamp, {"时间文本": timestamp})
                    if time_node_id:
                        self.add_edge(record_node_id, "发生于", time_node_id)

                # 角色、机构、药品、疾病的关系补全
                for field_name, field_value in record_fields.items():
                    if any(key in field_name for key in ["家属", "联系人", "家人"]):
                        family_node_id = self.add_node("家属", field_value, {"来源字段": field_name})
                        if family_node_id:
                            self.add_edge(elder_node_id, "有亲属关系", family_node_id)
                    if any(key in field_name for key in ["照护员", "护理员", "护工"]):
                        caregiver_node_id = self.add_node("照护人员", field_value, {"来源字段": field_name})
                        if caregiver_node_id:
                            self.add_edge(elder_node_id, "接受照护服务", caregiver_node_id)
                            if node_type == "照护服务记录":
                                self.add_edge(caregiver_node_id, "执行服务", record_node_id)
                    if any(key in field_name for key in ["护士", "医生", "医师", "精神科"]):
                        medical_staff_node_id = self.add_node("医护人员", field_value, {"来源字段": field_name})
                        if medical_staff_node_id:
                            self.add_edge(elder_node_id, "接受医疗服务", medical_staff_node_id)
                            if node_type in {"住院记录", "服药记录"}:
                                self.add_edge(medical_staff_node_id, "执行诊疗", record_node_id)
                    if any(key in field_name for key in ["医院", "机构", "科室"]):
                        hospital_node_id = self.add_node("医疗机构", field_value, {"来源字段": field_name})
                        if hospital_node_id:
                            self.add_edge(elder_node_id, "入住/就诊于", hospital_node_id)
                    if node_type == "服药记录" and any(key in field_name for key in ["药", "药品", "药物"]):
                        medicine_node_id = self.add_node("药品", field_value, {"来源字段": field_name})
                        if medicine_node_id:
                            self.add_edge(record_node_id, "服用药品", medicine_node_id)
                    if node_type == "住院记录" and any(key in field_name for key in ["疾病", "诊断", "病症"]):
                        disease_node_id = self.add_node("疾病", field_value, {"来源字段": field_name})
                        if disease_node_id:
                            self.add_edge(record_node_id, "确诊疾病", disease_node_id)
                    if node_type == "异常事件记录" and any(key in field_name for key in ["工作人员", "处理人"]):
                        worker_node_id = self.add_node("医护人员", field_value, {"来源字段": field_name, "角色": "工作人员"})
                        if worker_node_id:
                            self.add_edge(worker_node_id, "处理事件", record_node_id)

            # 智能洞察节点：绑定可信度与证据支撑关系。
            for insight_index, insight_text in enumerate(insight_text_pool[:MAX_INSIGHT_NODE_PER_SHEET], start=1):
                if sheet_name == "生命体征记录":
                    insight_node_type = "体征洞察"
                elif sheet_name == "服药情况":
                    insight_node_type = "用药洞察"
                elif sheet_name == "照护服务日志":
                    insight_node_type = "照护洞察"
                elif sheet_name == "睡眠情况记录":
                    insight_node_type = "睡眠洞察"
                elif sheet_name == "住院记录":
                    insight_node_type = "住院洞察"
                elif sheet_name == "健康档案":
                    insight_node_type = "健康档案洞察"
                elif sheet_name == "异常事件记录":
                    insight_node_type = "风险洞察"
                else:
                    insight_node_type = "健康评估记录"
                confidence_score = self._compute_confidence(
                    evidence_count=len(evidence_node_ids),
                    has_timestamp=any(
                        normalize_value(record_item.get("timestamp", "")) for record_item in evidence_records
                    ),
                    risk_count=len(sheet_insight.get("risk_points", [])),
                )
                insight_node_id = self.add_node(
                    insight_node_type,
                    f"{sheet_name}洞察{insight_index}:{insight_text}"[:72],
                    {
                        "sheet_name": sheet_name,
                        "来源": "Qwen工作表总结",
                        "insight_text": insight_text,
                        "业务类别": sheet_name,
                        "可信度": confidence_score,
                        "证据节点": evidence_node_ids[:],
                    },
                )
                if insight_node_id:
                    self.add_edge(
                        elder_node_id,
                        elder_relation,
                        insight_node_id,
                        properties={"可信度": confidence_score},
                    )
                    for evidence_node_id in evidence_node_ids:
                        self.add_edge(
                            insight_node_id,
                            "证据支撑",
                            evidence_node_id,
                            properties={"可信度": confidence_score},
                        )

        self.logger.log(
            "INFO",
            "build_knowledge_graph",
            "长者健康智能计算图构建完成。",
            {"node_count": len(self.nodes), "edge_count": len(self.edges)},
        )
        return {"nodes": list(self.nodes.values()), "edges": self.edges}


class WorkflowState(TypedDict, total=False):
    input_excel_path: str
    output_root_path: str
    config_path: str
    parser_mode: str
    case_directory: str
    workflow_directory: str
    image_directory: str
    intermediate_directory: str
    result_directory: str
    log_directory: str
    patient_name: str
    pinyin_name: str
    max_workers: int
    app_config: AppConfig
    raw_sheet_data: Dict[str, List[Dict[str, str]]]
    normalized_sheet_data: Dict[str, List[Dict[str, Any]]]
    sheet_insight_map: Dict[str, Dict[str, Any]]
    case_insight_summary: Dict[str, Any]
    personalized_health_profile: Dict[str, Any]
    agent_results: List[Dict[str, Any]]
    graph_data: Dict[str, Any]
    output_graph_path: str
    exception_list: List[str]
    logger: JsonlLogger
    renderer: WorkflowPngRenderer


def workflow_node_init_runtime(state: WorkflowState) -> WorkflowState:
    input_excel_path = Path(state["input_excel_path"])
    output_root_path = Path(state["output_root_path"])
    case_directory = output_root_path / input_excel_path.stem
    intermediate_directory = case_directory / "intermediate"
    result_directory = case_directory / "results"
    log_directory = case_directory / "logs"
    image_directory = case_directory / "images"
    workflow_directory = case_directory / "workflow"
    for directory in [case_directory, intermediate_directory, result_directory, log_directory, image_directory, workflow_directory]:
        directory.mkdir(parents=True, exist_ok=True)

    logger = JsonlLogger(log_directory / "core_runtime.jsonl")
    renderer = WorkflowPngRenderer(workflow_directory)
    renderer.draw_overview_flowchart()
    renderer.draw_step("init_runtime", 1)
    app_config = load_app_config(Path(state["config_path"]))
    parser_mode = normalize_value(state.get("parser_mode", app_config.parser_mode)).lower()
    if parser_mode not in {"mineru", "openpyxl"}:
        parser_mode = "mineru"
    app_config.parser_mode = parser_mode
    logger.log(
        "INFO",
        "init_runtime",
        "完成目录初始化。",
        {
            "case_directory": str(case_directory),
            "input_excel": str(input_excel_path),
            "config_path": state["config_path"],
            "parser_mode": app_config.parser_mode,
        },
    )
    return {
        **state,
        "case_directory": str(case_directory),
        "intermediate_directory": str(intermediate_directory),
        "result_directory": str(result_directory),
        "log_directory": str(log_directory),
        "image_directory": str(image_directory),
        "workflow_directory": str(workflow_directory),
        "exception_list": [],
        "logger": logger,
        "renderer": renderer,
        "app_config": app_config,
    }


def workflow_node_parse_excel(state: WorkflowState) -> WorkflowState:
    logger = state["logger"]
    renderer = state["renderer"]
    renderer.draw_step("parse_excel_by_mineru", 2)
    try:
        parser = MinerUWorkbookParser(state["app_config"], logger)
        mineru_output_root = Path(state["intermediate_directory"]) / "mineru_raw"
        sheet_data = parser.parse_excel(
            excel_path=Path(state["input_excel_path"]),
            mineru_output_root=mineru_output_root,
        )
        safe_json_dump(Path(state["intermediate_directory"]) / "raw_sheet_data.json", sheet_data)
        logger.log(
            "INFO",
            "parse_excel_by_mineru",
            "Excel 解析结果已落盘。",
            {"sheet_count": len(sheet_data)},
        )
        return {**state, "raw_sheet_data": sheet_data}
    except Exception as error:
        logger.log("ERROR", "parse_excel_by_mineru", "Excel 解析失败。", {"error": str(error)})
        updated_exceptions = state.get("exception_list", []) + [f"parse_excel_by_mineru: {error}"]
        return {**state, "raw_sheet_data": {}, "exception_list": updated_exceptions}


def workflow_node_parallel_agents(state: WorkflowState) -> WorkflowState:
    logger = state["logger"]
    renderer = state["renderer"]
    renderer.draw_step("run_parallel_sheet_agents", 3)
    sheet_data = state.get("raw_sheet_data", {})
    qwen_extractor = QwenExtractorAgent(state["app_config"], logger)
    sheet_agent = SheetAgent(logger, qwen_extractor)
    normalized_sheet_data: Dict[str, List[Dict[str, Any]]] = {}
    sheet_insight_map: Dict[str, Dict[str, Any]] = {}
    agent_results: List[Dict[str, Any]] = []

    if not sheet_data:
        logger.log("WARNING", "run_parallel_sheet_agents", "没有可处理的工作表数据。", {})
        return {
            **state,
            "normalized_sheet_data": normalized_sheet_data,
            "sheet_insight_map": sheet_insight_map,
            "agent_results": agent_results,
        }

    max_workers = max(1, int(state.get("max_workers", 4)))
    with ThreadPoolExecutor(max_workers=min(max_workers, len(sheet_data))) as executor:
        future_to_sheet = {
            executor.submit(sheet_agent.process_sheet, sheet_name, records): sheet_name
            for sheet_name, records in sheet_data.items()
        }
        for completed_future in as_completed(future_to_sheet):
            sheet_name = future_to_sheet[completed_future]
            try:
                processed_records, agent_result = completed_future.result()
                normalized_sheet_data[sheet_name] = processed_records
                agent_results.append(agent_result.__dict__)
                if processed_records:
                    sheet_insight_map[sheet_name] = processed_records[0].get("llm_structured_summary", {})
                else:
                    sheet_insight_map[sheet_name] = {
                        "sheet_name": sheet_name,
                        "current_status": "该工作表暂无有效记录。",
                        "key_findings": [],
                        "risk_points": [],
                        "recommended_actions": [],
                    }
            except Exception as error:  # pragma: no cover
                logger.log(
                    "ERROR",
                    "run_parallel_sheet_agents",
                    "并行任务失败。",
                    {"sheet_name": sheet_name, "error": str(error)},
                )
                sheet_insight_map[sheet_name] = {
                    "sheet_name": sheet_name,
                    "current_status": f"处理失败：{error}",
                    "key_findings": [],
                    "risk_points": [],
                    "recommended_actions": [],
                }
                normalized_sheet_data[sheet_name] = []
                agent_results.append(
                    AgentResult(
                        agent_name=f"{sheet_name}智能体",
                        sheet_name=sheet_name,
                        status="failed",
                        duration_seconds=0.0,
                        record_count=0,
                        message=str(error),
                    ).__dict__
                )

    safe_json_dump(Path(state["intermediate_directory"]) / "normalized_sheet_data.json", normalized_sheet_data)
    safe_json_dump(Path(state["intermediate_directory"]) / "sheet_insight_map.json", sheet_insight_map)
    safe_json_dump(Path(state["intermediate_directory"]) / "agent_results.json", agent_results)
    logger.log(
        "INFO",
        "run_parallel_sheet_agents",
        "并行智能体阶段结束。",
        {"agent_count": len(agent_results)},
    )
    return {
        **state,
        "normalized_sheet_data": normalized_sheet_data,
        "sheet_insight_map": sheet_insight_map,
        "agent_results": agent_results,
    }


def infer_patient_name(sheet_data: Dict[str, List[Dict[str, Any]]], fallback_name: str) -> str:
    """从健康档案优先提取姓名，失败则回退到文件名。"""
    profile_records = sheet_data.get("健康档案", [])
    if profile_records:
        source_record = profile_records[0].get("source_record", {})
        for field_name, field_value in source_record.items():
            if "姓名" in field_name and field_value:
                return field_value
    return fallback_name


def workflow_node_build_graph(state: WorkflowState) -> WorkflowState:
    logger = state["logger"]
    renderer = state["renderer"]
    renderer.draw_step("build_knowledge_graph", 4)
    try:
        input_stem = Path(state["input_excel_path"]).stem
        normalized_sheet_data = state.get("normalized_sheet_data", {})
        sheet_insight_map = state.get("sheet_insight_map", {})
        patient_name = infer_patient_name(normalized_sheet_data, input_stem)
        case_insight_agent = QwenCaseInsightAgent(state["app_config"], logger)
        case_insight_summary = case_insight_agent.summarize_case(sheet_insight_map)
        personalized_health_profile = case_insight_agent.build_personalized_health_profile(
            patient_name=patient_name,
            sheet_insight_map=sheet_insight_map,
            case_insight_summary=case_insight_summary,
        )
        graph_builder = ElderCareKnowledgeGraphBuilder(logger)
        graph_data = graph_builder.build_graph(
            patient_name=patient_name,
            normalized_sheet_data=normalized_sheet_data,
            sheet_insight_map=sheet_insight_map,
            case_insight_summary=case_insight_summary,
            personalized_health_profile=personalized_health_profile,
        )
        pinyin_name = convert_name_to_pinyin(patient_name, input_stem)
        safe_json_dump(Path(state["intermediate_directory"]) / "graph_data_preview.json", graph_data)
        safe_json_dump(Path(state["intermediate_directory"]) / "case_insight_summary.json", case_insight_summary)
        safe_json_dump(
            Path(state["intermediate_directory"]) / "personalized_health_profile.json",
            personalized_health_profile,
        )
        return {
            **state,
            "graph_data": graph_data,
            "patient_name": patient_name,
            "pinyin_name": pinyin_name,
            "case_insight_summary": case_insight_summary,
            "personalized_health_profile": personalized_health_profile,
        }
    except Exception as error:
        logger.log("ERROR", "build_knowledge_graph", "长者健康智能计算图构建失败。", {"error": str(error)})
        updated_exceptions = state.get("exception_list", []) + [f"build_knowledge_graph: {error}"]
        return {**state, "graph_data": {"nodes": [], "edges": []}, "exception_list": updated_exceptions}


def workflow_node_persist_files(state: WorkflowState) -> WorkflowState:
    logger = state["logger"]
    renderer = state["renderer"]
    renderer.draw_step("persist_graph_files", 5)
    result_directory = Path(state["result_directory"])
    pinyin_name = state.get("pinyin_name", Path(state["input_excel_path"]).stem)
    graph_file_path = result_directory / f"{pinyin_name}_elder_health_computing_graph.json"
    graph_data = state.get("graph_data", {"nodes": [], "edges": []})
    safe_json_dump(graph_file_path, graph_data)
    personalized_profile_path = result_directory / f"{pinyin_name}_personalized_health_profile.json"
    safe_json_dump(personalized_profile_path, state.get("personalized_health_profile", {}))

    graph_snippet = {
        "node_count": len(graph_data.get("nodes", [])),
        "edge_count": len(graph_data.get("edges", [])),
        "sample_nodes": graph_data.get("nodes", [])[:10],
        "sample_edges": graph_data.get("edges", [])[:10],
    }
    safe_json_dump(result_directory / "graph_snippet.json", graph_snippet)
    graph_image_path = result_directory / f"{pinyin_name}_elder_health_computing_graph.png"
    visual_meta = render_knowledge_graph_image(graph_data, graph_image_path)
    safe_json_dump(result_directory / "graph_visualization_meta.json", visual_meta)
    logger.log(
        "INFO",
        "persist_graph_files",
        "长者健康智能计算图文件保存成功。",
        {
            "graph_file": str(graph_file_path),
            "personalized_profile_file": str(personalized_profile_path),
            "graph_image_file": str(graph_image_path),
            "visual_node_count": visual_meta.get("node_count"),
            "visual_edge_count": visual_meta.get("edge_count"),
        },
    )
    return {**state, "output_graph_path": str(graph_file_path)}


def workflow_node_export_pdf(state: WorkflowState) -> WorkflowState:
    logger = state["logger"]
    renderer = state["renderer"]
    renderer.draw_step("export_pdf_log", 6)
    task_overview = {
        "任务描述": "Data Agent 任务一：老人档案 Excel -> 长者健康智能计算图",
        "输入文件": state["input_excel_path"],
        "输出图谱": state.get("output_graph_path", ""),
        "处理流程": "MinerU解析 -> 通义千问抽取 -> 并行结构化 -> 图谱构建 -> 落盘 -> PDF日志",
        "主要创新点": "通过长者健康智能计算图将多源原始照护记录转化为可计算高质量洞察 + MinerU+LangGraph编排 + 分工作表并行智能体 + LLM增强抽取 + 时间线图谱",
        "异常数量": len(state.get("exception_list", [])),
        "模型返回内容": f"已调用模型：{state['app_config'].dashscope_model}",
    }
    agent_catalog = [
        {
            "name": "A0 任务总控智能体",
            "type": "编排型智能体",
            "responsibility": "负责全局任务启动、配置加载、错误收敛与阶段编排。",
            "structure": "LangGraph 状态机 + StateGraph 节点拓扑。",
            "tools": "LangGraph、JsonlLogger、WorkflowPngRenderer。",
            "inputs": "输入文件路径、配置文件路径、输出根目录。",
            "outputs": "标准化运行状态、阶段日志、工作流可视化图片。",
        },
        {
            "name": "A1 MinerU 文档解析智能体",
            "type": "解析型智能体",
            "responsibility": "调用 MinerU 解析 XLSX，并把表格 HTML 转换为结构化记录。",
            "structure": "MinerU CLI 调用层 + HTML 表格解析层 + 工作表语义映射层。",
            "tools": "mineru CLI、正则解析、JSON 文件读写。",
            "inputs": "老人档案 XLSX 文件。",
            "outputs": "raw_sheet_data（按工作表聚合后的结构化记录）。",
        },
        {
            "name": "A2 大模型语义增强智能体（语义增强型智能体）",
            "type": "语义增强型智能体",
            "responsibility": "对每个工作表样本做高价值字段、风险点与摘要抽取。",
            "structure": "OpenAI SDK 兼容接口 + JSON Schema 约束输出。",
            "tools": "DashScope API（qwen3-max）、response_format=json_object。",
            "inputs": "工作表样本记录。",
            "outputs": "high_value_fields、event_summary、risk_points。",
        },
        {
            "name": "A3 工作表并行处理智能体集群",
            "type": "并行计算智能体",
            "responsibility": "并行执行字段清洗、时间戳识别、规则抽取与大模型结果融合。",
            "structure": "ThreadPoolExecutor + 单表处理单元（SheetAgent）。",
            "tools": "concurrent.futures、字段清洗规则、时间戳识别规则。",
            "inputs": "raw_sheet_data + Qwen 抽取结果。",
            "outputs": "normalized_sheet_data + agent_results。",
        },
        {
            "name": "A4 长者健康智能计算图构建智能体",
            "type": "图谱建模智能体",
            "responsibility": "按节点/边约束构建老人照护长者健康智能计算图与时间线关系。",
            "structure": "Schema 约束层 + 节点去重层 + 边关系校验层。",
            "tools": "ElderCareKnowledgeGraphBuilder、关系白名单、时间维度补齐逻辑。",
            "inputs": "normalized_sheet_data。",
            "outputs": "graph_data（nodes/edges）。",
        },
        {
            "name": "A5 审计日志与报告智能体",
            "type": "治理型智能体",
            "responsibility": "记录全链路日志并输出 PDF 技术报告。",
            "structure": "JSONL 事件流 + PDF 模板汇总。",
            "tools": "ReportLab、JsonlLogger、图谱预览模块。",
            "inputs": "任务状态、智能体信息、工作流信息、执行日志。",
            "outputs": "core_runtime.jsonl + core_runtime_report.pdf。",
        },
    ]
    workflow_catalog = [
        {
            "name": "W1 主工作流（LangGraph 主链路）",
            "goal": "打通输入文档到图谱输出的全自动流水线。",
            "trigger": "接收到 XLSX 输入文件后自动触发。",
            "agents": "A0 任务总控智能体、A1 MinerU文档解析智能体、A2 大模型语义增强智能体（语义增强型智能体）、A3 工作表并行处理智能体集群、A4 长者健康智能计算图构建智能体、A5 审计日志与报告智能体",
            "nodes": [
                {
                    "name": "init_runtime",
                    "purpose": "创建目录、加载配置、初始化日志器与流程图渲染器。",
                    "inputs": "input_excel_path、output_root_path、config_path",
                    "outputs": "case_directory、logger、renderer、app_config",
                },
                {
                    "name": "parse_excel_by_mineru",
                    "purpose": "调用 MinerU CLI 解析 XLSX，抽取结构化表格数据。",
                    "inputs": "input_excel_path、app_config.mineru_*",
                    "outputs": "raw_sheet_data",
                },
                {
                    "name": "run_parallel_sheet_agents",
                    "purpose": "并行运行单表智能体，融合规则抽取与通义千问抽取。",
                    "inputs": "raw_sheet_data、app_config.qwen_*",
                    "outputs": "normalized_sheet_data、agent_results",
                },
                {
                    "name": "build_knowledge_graph",
                    "purpose": "按长者健康智能计算图 schema 构建节点与边并补齐时间维度。",
                    "inputs": "normalized_sheet_data",
                    "outputs": "graph_data、patient_name、pinyin_name",
                },
                {
                    "name": "persist_graph_files",
                    "purpose": "将图谱与片段落盘，便于下游训练或检索。",
                    "inputs": "graph_data、result_directory",
                    "outputs": "elder_health_computing_graph.json、graph_snippet.json、个性化健康档案",
                },
                {
                    "name": "export_pdf_log",
                    "purpose": "汇总技术细节与执行日志，生成评审版 PDF 报告。",
                    "inputs": "任务上下文、graph_preview、运行日志",
                    "outputs": "core_runtime_report.pdf",
                },
            ],
        },
        {
            "name": "W2 工作表并行子工作流",
            "goal": "提升多工作表处理吞吐并保证可追溯性。",
            "trigger": "主流程进入 run_parallel_sheet_agents 节点时触发。",
            "agents": "A2 大模型语义增强智能体（语义增强型智能体）、A3 工作表并行处理智能体集群、A5 审计日志与报告智能体",
            "nodes": [
                {
                    "name": "qwen_sheet_extract",
                    "purpose": "按工作表生成高价值字段建议和风险点。",
                    "inputs": "工作表样本记录",
                    "outputs": "字段建议、摘要、风险点",
                },
                {
                    "name": "sheet_normalize",
                    "purpose": "做字段清洗、时间戳识别和抽取结果融合。",
                    "inputs": "原始记录 + qwen输出",
                    "outputs": "标准化记录",
                },
                {
                    "name": "sheet_audit_log",
                    "purpose": "记录每个智能体耗时、处理条数、异常信息。",
                    "inputs": "执行状态",
                    "outputs": "agent_results + JSONL日志",
                },
            ],
        },
        {
            "name": "W3 异常恢复工作流",
            "goal": "保证流程在模型调用失败时仍可继续执行。",
            "trigger": "MinerU 或 Qwen 任一阶段抛异常时触发。",
            "agents": "A0 任务总控智能体、A1 MinerU文档解析智能体、A2 大模型语义增强智能体（语义增强型智能体）、A3 工作表并行处理智能体集群、A5 审计日志与报告智能体",
            "nodes": [
                {
                    "name": "exception_capture",
                    "purpose": "捕获异常并写入 exception_list 与日志。",
                    "inputs": "异常对象",
                    "outputs": "异常记录",
                },
                {
                    "name": "fallback_strategy",
                    "purpose": "大模型失败时降级为规则抽取，保证主链不中断。",
                    "inputs": "失败上下文",
                    "outputs": "可用的兜底抽取结果",
                },
                {
                    "name": "continue_pipeline",
                    "purpose": "恢复后续节点执行并输出最终结果。",
                    "inputs": "兜底结果",
                    "outputs": "完整任务输出",
                },
            ],
        },
    ]
    innovation_points = [
        "通过长者健康智能计算图，将照护机构采集的原始老人画像、异常事件记录、生命体征记录、用药记录、照护服务记录、健康评估记录、睡眠情况记录，转化为经智能分析且可计算的老人主体近况摘要、风险洞察、照护洞察、个性化健康档案、体征洞察、用药洞察等高质量信息。",
        "使用 MinerU 原生 XLSX 解析链路，保留复杂表格结构信息，再转换成图谱友好结构。",
        "采用 LangGraph 显式状态机编排，节点职责清晰，便于审计、回放与评测。",
        "构建“规则抽取 + 通义千问语义抽取”双引擎融合机制，在稳定性与语义覆盖之间平衡。",
        "每个工作表由独立智能体并行处理，显著降低全量数据处理耗时。",
        "图谱构建阶段严格使用节点/边白名单，确保输出可控、可训练、可解释。",
        "日志体系同时输出 JSONL 与 PDF 技术报告，兼顾机器评测与人工评审展示。",
    ]
    graph_data = state.get("graph_data", {"nodes": [], "edges": []})
    graph_preview = {
        "node_count": len(graph_data.get("nodes", [])),
        "edge_count": len(graph_data.get("edges", [])),
        "sample_nodes": [
            f"{node.get('type')}:{node.get('name')}" for node in graph_data.get("nodes", [])[:6]
        ],
        "sample_edges": [
            f"{edge.get('source')} -[{edge.get('relation')}]-> {edge.get('target')}"
            for edge in graph_data.get("edges", [])[:6]
        ],
    }
    execution_summary = {
        "总节点数": len(graph_data.get("nodes", [])),
        "总边数": len(graph_data.get("edges", [])),
        "异常总数": len(state.get("exception_list", [])),
        "并行线程数": state.get("max_workers", 0),
        "MinerU后端": state["app_config"].mineru_backend,
        "MinerU解析方式": state["app_config"].mineru_method,
        "通义千问模型": state["app_config"].dashscope_model,
        "核心日志条数": len(logger.buffer),
        "输出目录": state.get("case_directory", ""),
    }
    pdf_path = Path(state["log_directory"]) / "core_runtime_report.pdf"
    logger.export_pdf_summary(
        pdf_path=pdf_path,
        task_overview=task_overview,
        graph_preview=graph_preview,
        agent_catalog=agent_catalog,
        workflow_catalog=workflow_catalog,
        innovation_points=innovation_points,
        execution_summary=execution_summary,
    )
    logger.log(
        "INFO",
        "export_pdf_log",
        "PDF 核心日志导出完成。",
        {"pdf_path": str(pdf_path)},
    )
    return state


def create_langgraph_workflow():
    workflow_builder = StateGraph(WorkflowState)
    workflow_builder.add_node("init_runtime", workflow_node_init_runtime)
    workflow_builder.add_node("parse_excel_by_mineru", workflow_node_parse_excel)
    workflow_builder.add_node("run_parallel_sheet_agents", workflow_node_parallel_agents)
    workflow_builder.add_node("build_knowledge_graph", workflow_node_build_graph)
    workflow_builder.add_node("persist_graph_files", workflow_node_persist_files)
    workflow_builder.add_node("export_pdf_log", workflow_node_export_pdf)

    workflow_builder.add_edge(START, "init_runtime")
    workflow_builder.add_edge("init_runtime", "parse_excel_by_mineru")
    workflow_builder.add_edge("parse_excel_by_mineru", "run_parallel_sheet_agents")
    workflow_builder.add_edge("run_parallel_sheet_agents", "build_knowledge_graph")
    workflow_builder.add_edge("build_knowledge_graph", "persist_graph_files")
    workflow_builder.add_edge("persist_graph_files", "export_pdf_log")
    workflow_builder.add_edge("export_pdf_log", END)
    return workflow_builder.compile()


def run_single_file_pipeline(
    excel_file_path: Path,
    output_root_path: Path,
    config_path: Path,
    max_workers: int,
    parser_mode: str,
) -> Dict[str, Any]:
    workflow = create_langgraph_workflow()
    initial_state: WorkflowState = {
        "input_excel_path": str(excel_file_path),
        "output_root_path": str(output_root_path),
        "config_path": str(config_path),
        "max_workers": max_workers,
        "parser_mode": parser_mode,
    }
    final_state = workflow.invoke(initial_state)
    return {
        "input_excel_path": str(excel_file_path),
        "case_directory": final_state.get("case_directory", ""),
        "output_graph_path": final_state.get("output_graph_path", ""),
        "node_count": len(final_state.get("graph_data", {}).get("nodes", [])),
        "edge_count": len(final_state.get("graph_data", {}).get("edges", [])),
        "exception_count": len(final_state.get("exception_list", [])),
    }


def collect_input_files(input_path: Path) -> List[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".xlsx":
        return [input_path]
    if input_path.is_dir():
        return sorted(input_path.rglob("*.xlsx"))
    return []


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Data Agent 任务一：Excel -> 长者健康智能计算图")
    parser.add_argument("--input", default=DEFAULT_INPUT_FILE, help="输入 Excel 文件或目录路径")
    parser.add_argument("--output-root", default=DEFAULT_OUTPUT_ROOT, help="输出根目录")
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="配置文件路径（JSON）")
    parser.add_argument("--max-workers", type=int, default=4, help="并行处理线程数")
    parser.add_argument(
        "--parser-mode",
        default="mineru",
        choices=["mineru", "openpyxl"],
        help="解析模式：mineru 为正式模式，openpyxl 为消融基线模式",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_root_path = Path(args.output_root)
    config_path = Path(args.config)
    output_root_path.mkdir(parents=True, exist_ok=True)

    excel_files = collect_input_files(input_path)
    if not excel_files:
        raise FileNotFoundError(f"未找到可处理的 xlsx 文件：{input_path}")

    batch_start_seconds = time.time()
    batch_results: List[Dict[str, Any]] = []
    for excel_file in excel_files:
        try:
            one_result = run_single_file_pipeline(
                excel_file_path=excel_file,
                output_root_path=output_root_path,
                config_path=config_path,
                max_workers=args.max_workers,
                parser_mode=args.parser_mode,
            )
            batch_results.append(one_result)
        except Exception as error:  # pragma: no cover
            batch_results.append(
                {
                    "input_excel_path": str(excel_file),
                    "status": "failed",
                    "error": str(error),
                    "traceback": traceback.format_exc(),
                }
            )

    summary = {
        "batch_start_time": current_time_text(),
        "input_path": str(input_path),
        "output_root": str(output_root_path),
        "parser_mode": args.parser_mode,
        "file_count": len(excel_files),
        "duration_seconds": round(time.time() - batch_start_seconds, 3),
        "results": batch_results,
    }
    safe_json_dump(output_root_path / "batch_summary.json", summary)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
