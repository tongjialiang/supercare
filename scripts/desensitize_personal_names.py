#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""全项目人名脱敏：文本替换、路径重命名、Excel 单元格替换。"""

from __future__ import annotations

import re
import sys
from pathlib import Path

PROJECT_ROOT = Path("/srv/supercare")

# 按长度降序，避免短串误替换
TEXT_REPLACEMENTS: list[tuple[str, str]] = [
    ("陈萍忠女士", "陈女士"),
    ("chenzhongping_CareCase", "ms_chen_CareCase"),
    ("Chenzhongping_CareCase", "ms_chen_CareCase"),
    ("CHENZHONGPING_CARECASE", "MS_CHEN_CARECASE"),
    ("陈萍忠", "陈女士"),
    ("陈平忠", "陈女士"),
    ("chenzhongping", "ms_chen"),
    ("Chenzhongping", "ms_chen"),
    ("CHENZHONGPING", "MS_CHEN"),
    # 队列其他病例（DataSource 目录名）
    ("凌宝玉", "凌女士"),
    ("宋树成", "宋先生"),
    ("庄福舜", "庄先生"),
    ("梁徳馨", "梁女士"),
    ("殷春花", "殷女士"),
    ("江甦", "江女士"),
    ("裘爱芳", "裘女士"),
    ("陈如琴", "陈阿姨"),
    ("黄友娣", "黄女士"),
    # 照护人员
    ("戴梦遥", "护理员甲"),
    ("王春霞", "护理员乙"),
    ("彭蓉", "护士甲"),
]

# 修正可能出现的重复称谓
POST_FIXES: list[tuple[str, str]] = [
    ("陈女士女士", "陈女士"),
    ("陈陈平忠", "陈女士"),  # 历史笔误残留
    ("陈陈女士", "陈女士"),
]

TEXT_EXTENSIONS = {
    ".py", ".md", ".txt", ".json", ".jsonl", ".log", ".yaml", ".yml",
    ".ini", ".cfg", ".toml", ".xml", ".html", ".csv", ".tsv", ".rst",
}

SKIP_DIR_NAMES = {
    ".git", ".cursor", "__pycache__", "node_modules", ".venv", "venv",
}


def apply_replacements(text: str) -> str:
    for old, new in TEXT_REPLACEMENTS:
        text = text.replace(old, new)
    for old, new in POST_FIXES:
        text = text.replace(old, new)
    return text


def should_process_file(path: Path) -> bool:
    if path.suffix.lower() in TEXT_EXTENSIONS:
        return True
    # 无扩展名但可能是脚本
    if path.suffix == "" and path.is_file() and path.stat().st_size < 2_000_000:
        try:
            sample = path.read_bytes()[:4096]
            if b"\x00" in sample:
                return False
            sample.decode("utf-8")
            return True
        except (UnicodeDecodeError, OSError):
            return False
    return False


def process_text_files(root: Path) -> int:
    changed = 0
    for path in root.rglob("*"):
        if any(part in SKIP_DIR_NAMES for part in path.parts):
            continue
        if not path.is_file() or not should_process_file(path):
            continue
        if path.name == "desensitize_personal_names.py":
            continue
        try:
            original = path.read_text(encoding="utf-8")
        except (UnicodeDecodeError, OSError):
            continue
        updated = apply_replacements(original)
        if updated != original:
            path.write_text(updated, encoding="utf-8")
            changed += 1
    return changed


def collect_rename_targets(root: Path) -> list[Path]:
    targets: list[Path] = []
    for path in root.rglob("*"):
        if any(part in SKIP_DIR_NAMES for part in path.parts):
            continue
        name = path.name
        if any(old in name for old, _ in TEXT_REPLACEMENTS):
            targets.append(path)
    # 最长路径先重命名，避免父路径先改导致子路径找不到
    return sorted(targets, key=lambda p: len(str(p)), reverse=True)


def rename_paths(root: Path) -> int:
    renamed = 0
    for path in collect_rename_targets(root):
        new_name = apply_replacements(path.name)
        if new_name == path.name:
            continue
        new_path = path.with_name(new_name)
        if new_path.exists():
            print(f"跳过（目标已存在）: {path} -> {new_path}", file=sys.stderr)
            continue
        path.rename(new_path)
        renamed += 1
    return renamed


def process_xlsx_files(root: Path) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("未安装 openpyxl，跳过 Excel 内嵌文本脱敏", file=sys.stderr)
        return 0

    changed = 0
    for path in root.rglob("*.xlsx"):
        if any(part in SKIP_DIR_NAMES for part in path.parts):
            continue
        try:
            workbook = load_workbook(path)
        except Exception as exc:
            print(f"无法打开 Excel: {path} ({exc})", file=sys.stderr)
            continue
        file_changed = False
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str):
                        new_value = apply_replacements(value)
                        if new_value != value:
                            cell.value = new_value
                            file_changed = True
        if file_changed:
            workbook.save(path)
            changed += 1
    return changed


def main() -> None:
    print("1/4 文本文件内容脱敏...")
    text_changed = process_text_files(PROJECT_ROOT)
    print(f"   已更新 {text_changed} 个文本文件")

    print("2/4 路径与文件名脱敏...")
    path_renamed = rename_paths(PROJECT_ROOT)
    print(f"   已重命名 {path_renamed} 个路径")

    print("3/4 Excel 单元格脱敏...")
    xlsx_changed = process_xlsx_files(PROJECT_ROOT)
    print(f"   已更新 {xlsx_changed} 个 Excel 文件")

    print("4/4 校验残留...")
    remaining = 0
    for old, _ in TEXT_REPLACEMENTS:
        if old in ("陈萍忠", "陈平忠", "chenzhongping", "chenzhongping_CareCase"):
            pass  # 下面统一检查
    check_patterns = ["陈萍忠", "陈平忠", "chenzhongping", "戴梦遥", "王春霞", "彭蓉"]
    for pattern in check_patterns:
        count = 0
        for path in PROJECT_ROOT.rglob("*"):
            if any(part in SKIP_DIR_NAMES for part in path.parts):
                continue
            if not path.is_file() or not should_process_file(path):
                continue
            try:
                if pattern in path.read_text(encoding="utf-8"):
                    count += 1
            except (UnicodeDecodeError, OSError):
                continue
        if count:
            print(f"   警告: 文本中仍含「{pattern}」的文件数: {count}")
            remaining += count
    if remaining == 0:
        print("   主要人名已无文本残留")
    print("完成。")


if __name__ == "__main__":
    main()
