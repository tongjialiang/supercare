# -*- coding: utf-8 -*-
"""从 DataSource 病例 Excel 提取老年健康生物标志物。"""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook

from staging.biomarker_catalog import BIOMARKER_DEFINITIONS

DATASOURCE_ROOT = Path("/srv/supercare/DataSource")
DEFAULT_EXCEL_OUTPUT = Path("/srv/supercare/task-agent/output/a0_老年健康生物标志物矩阵.xlsx")

_NUMERIC_PATTERN = re.compile(r"(-?\d+(?:\.\d+)?)")


def _safe_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _parse_number(text: str) -> Optional[float]:
    match = _NUMERIC_PATTERN.search(_safe_text(text))
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def _parse_date_key(value: Any) -> str:
    text = _safe_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:19] if " " in text else text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text[:10]


def _discover_case_excels(data_root: Path = DATASOURCE_ROOT) -> List[Tuple[str, Path]]:
    """返回 (病例文件夹名, excel路径) 列表。"""
    cases: List[Tuple[str, Path]] = []
    if not data_root.is_dir():
        return cases
    for folder in sorted(data_root.iterdir()):
        if not folder.is_dir():
            continue
        candidates = sorted(folder.glob("*CareCase*.xlsx"))
        if not candidates:
            continue
        # 优先非 standardized 的原始表（信息更全），否则用 standardized
        raw = [path for path in candidates if "standardized" not in path.name.lower()]
        excel_path = raw[0] if raw else candidates[0]
        cases.append((folder.name, excel_path))
    return cases


def _case_id_from_path(excel_path: Path, folder_name: str) -> str:
    stem = excel_path.stem.replace("_CareCase", "").replace("_standardized", "")
    return stem or folder_name


def _sum_npi_session(rows: List[Tuple[Any, ...]]) -> Optional[float]:
    """单次 NPI 评估：累加可解析为数值的答案（严重度/频度项）。"""
    total = 0.0
    counted = 0
    for _, _, question, answer, _ in rows:
        question_text = _safe_text(question)
        if "苦恼" in question_text and "程度" in question_text:
            continue
        if "总分" in question_text:
            score = _parse_number(_safe_text(answer))
            if score is not None:
                return score
        score = _parse_number(_safe_text(answer))
        if score is None:
            continue
        if score > 12:
            continue
        total += score
        counted += 1
    if counted == 0:
        return None
    return round(total, 2)


def _sum_barthel_session(rows: List[Tuple[Any, ...]]) -> Optional[float]:
    total = 0.0
    for _, _, _, answer, _ in rows:
        score = _parse_number(_safe_text(answer))
        if score is not None and score <= 20:
            total += score
    return round(total, 2) if total > 0 else None


def _extract_mmse_session(rows: List[Tuple[Any, ...]]) -> Optional[float]:
    for _, _, question, answer, _ in rows:
        if "总分" in _safe_text(question) or "合计" in _safe_text(question):
            score = _parse_number(_safe_text(answer))
            if score is not None:
                return round(score, 2)
    total = 0.0
    for _, _, _, answer, _ in rows:
        score = _parse_number(_safe_text(answer))
        if score is not None and score <= 1:
            total += score
    return round(total, 2) if total > 0 else None


def _load_assessment_snapshots(excel_path: Path) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "健康评估记录" not in workbook.sheetnames:
        workbook.close()
        return []
    worksheet = workbook["健康评估记录"]
    grouped: Dict[Tuple[str, str], List[Tuple[Any, ...]]] = defaultdict(list)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        assess_date = _parse_date_key(row[0])
        assess_name = _safe_text(row[1])
        grouped[(assess_date, assess_name)].append(row)

    snapshots: List[Dict[str, Any]] = []
    for (assess_date, assess_name), rows in grouped.items():
        record: Dict[str, Any] = {"date": assess_date, "assessment_name": assess_name}
        if "NPI" in assess_name:
            record["npi_total"] = _sum_npi_session(rows)
        elif assess_name == "Barthel":
            record["barthel_total"] = _sum_barthel_session(rows)
        elif "MMSE" in assess_name:
            record["mmse_total"] = _extract_mmse_session(rows)
        elif "MOCA" in assess_name.upper() or "MoCA" in assess_name:
            moca_score = _extract_mmse_session(rows)
            if moca_score is not None:
                record["moca_total"] = moca_score
                record["moca_inverse"] = round(100 - float(moca_score), 2)
        else:
            continue
        snapshots.append(record)
    workbook.close()
    snapshots.sort(key=lambda item: item.get("date", ""))
    return snapshots


def _monthly_vital_means(excel_path: Path) -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "生命体征记录" not in workbook.sheetnames:
        workbook.close()
        return []
    worksheet = workbook["生命体征记录"]
    buckets: Dict[Tuple[str, str], List[float]] = defaultdict(list)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        month_key = _parse_date_key(row[0])[:7]
        vital_type = _safe_text(row[1])
        value = _parse_number(_safe_text(row[2]))
        if not month_key or value is None:
            continue
        if vital_type in ("收缩压", "舒张压", "心率", "血氧"):
            buckets[(month_key, vital_type)].append(value)
    workbook.close()

    monthly: Dict[str, Dict[str, Any]] = {}
    for (month_key, vital_type), values in buckets.items():
        if not values:
            continue
        monthly.setdefault(month_key, {"month": month_key})
        monthly[month_key][vital_type] = round(sum(values) / len(values), 2)
    result = list(monthly.values())
    result.sort(key=lambda item: item.get("month", ""))
    return result


def extract_single_case(excel_path: Path, folder_name: str = "") -> Dict[str, Any]:
    """提取单病例生物标志物时间序列。"""
    case_id = _case_id_from_path(excel_path, folder_name)
    assessments = _load_assessment_snapshots(excel_path)
    vitals = _monthly_vital_means(excel_path)

    longitudinal: List[Dict[str, Any]] = []
    for item in assessments:
        row: Dict[str, Any] = {
            "case_id": case_id,
            "folder_name": folder_name or case_id,
            "time_key": item.get("date", ""),
            "source": "健康评估记录",
        }
        if item.get("npi_total") is not None:
            row["npi_total"] = item["npi_total"]
        if item.get("barthel_total") is not None:
            row["barthel_total"] = item["barthel_total"]
            row["adl_dependence"] = round(100 - float(item["barthel_total"]), 2)
        if item.get("mmse_total") is not None:
            row["mmse_total"] = item["mmse_total"]
            row["cognitive_inverse"] = round(100 - float(item["mmse_total"]), 2)
        if item.get("moca_total") is not None:
            row["moca_total"] = item["moca_total"]
        if item.get("moca_inverse") is not None:
            row["moca_inverse"] = item["moca_inverse"]
        longitudinal.append(row)

    for vital in vitals:
        longitudinal.append(
            {
                "case_id": case_id,
                "folder_name": folder_name or case_id,
                "time_key": vital.get("month", ""),
                "source": "生命体征记录",
                "systolic_bp": vital.get("收缩压"),
                "diastolic_bp": vital.get("舒张压"),
                "heart_rate": vital.get("心率"),
                "spo2": vital.get("血氧"),
            }
        )

    latest = _build_latest_snapshot(longitudinal, case_id, folder_name or case_id, excel_path)
    return {
        "case_id": case_id,
        "folder_name": folder_name or case_id,
        "excel_path": str(excel_path),
        "longitudinal": longitudinal,
        "latest_snapshot": latest,
    }


def _build_latest_snapshot(
    longitudinal: List[Dict[str, Any]],
    case_id: str,
    folder_name: str,
    excel_path: Path,
) -> Dict[str, Any]:
    """合并各标志物最近一次有效值（评估与体征分轨，避免月份键覆盖评估日期）。"""
    snapshot: Dict[str, Any] = {
        "case_id": case_id,
        "folder_name": folder_name,
        "excel_path": str(excel_path),
    }
    assess_keys = (
        "npi_total",
        "barthel_total",
        "adl_dependence",
        "mmse_total",
        "cognitive_inverse",
        "moca_total",
        "moca_inverse",
    )
    vital_keys = ("systolic_bp", "diastolic_bp", "heart_rate", "spo2")
    assess_rows = sorted(
        [row for row in longitudinal if row.get("source") == "健康评估记录"],
        key=lambda item: item.get("time_key", ""),
        reverse=True,
    )
    vital_rows = sorted(
        [row for row in longitudinal if row.get("source") == "生命体征记录"],
        key=lambda item: item.get("time_key", ""),
        reverse=True,
    )
    for row in assess_rows:
        for key in assess_keys:
            if key not in snapshot and row.get(key) is not None:
                snapshot[key] = row[key]
        if all(key in snapshot for key in assess_keys):
            break
    for row in vital_rows:
        for key in vital_keys:
            if key not in snapshot and row.get(key) is not None:
                snapshot[key] = row[key]
    if assess_rows:
        snapshot["last_assessment_date"] = assess_rows[0].get("time_key", "")
    if vital_rows:
        snapshot["last_vital_month"] = vital_rows[0].get("time_key", "")
    return snapshot


def extract_cohort_biomarkers(data_root: Path = DATASOURCE_ROOT) -> Dict[str, Any]:
    """提取 DataSource 全队列病例。"""
    cases = _discover_case_excels(data_root)
    cohort: List[Dict[str, Any]] = []
    all_longitudinal: List[Dict[str, Any]] = []
    for folder_name, excel_path in cases:
        case_payload = extract_single_case(excel_path, folder_name)
        cohort.append(case_payload)
        all_longitudinal.extend(case_payload["longitudinal"])
    return {
        "extracted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "data_root": str(data_root),
        "case_count": len(cohort),
        "cases": cohort,
        "longitudinal_rows": all_longitudinal,
        "biomarker_definitions": BIOMARKER_DEFINITIONS,
    }


def write_biomarker_excel(cohort_payload: Dict[str, Any], output_path: Path = DEFAULT_EXCEL_OUTPUT) -> Path:
    """写入生物标志物 Excel（明细 + 队列最新截面）。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    sheet_detail = workbook.active
    sheet_detail.title = "生物标志物明细"
    detail_headers = [
        "病例ID",
        "文件夹名",
        "时间点",
        "数据来源",
        "NPI总分",
        "Barthel总分",
        "ADL依赖度",
        "MMSE总分",
        "认知逆指标",
        "收缩压月均",
        "舒张压月均",
        "心率月均",
        "血氧月均",
    ]
    sheet_detail.append(detail_headers)
    for case in cohort_payload.get("cases", []):
        for row in case.get("longitudinal", []):
            sheet_detail.append(
                [
                    row.get("case_id", ""),
                    row.get("folder_name", ""),
                    row.get("time_key", ""),
                    row.get("source", ""),
                    row.get("npi_total"),
                    row.get("barthel_total"),
                    row.get("adl_dependence"),
                    row.get("mmse_total"),
                    row.get("cognitive_inverse"),
                    row.get("systolic_bp"),
                    row.get("diastolic_bp"),
                    row.get("heart_rate"),
                    row.get("spo2"),
                ]
            )

    sheet_cohort = workbook.create_sheet("队列最新截面")
    cohort_headers = [
        "病例ID",
        "文件夹名",
        "最近评估/体征时间",
        "NPI总分",
        "Barthel总分",
        "ADL依赖度",
        "MMSE总分",
        "认知逆指标",
        "收缩压",
        "心率",
        "血氧",
        "Excel路径",
    ]
    sheet_cohort.append(cohort_headers)
    for case in cohort_payload.get("cases", []):
        snap = case.get("latest_snapshot", {})
        sheet_cohort.append(
            [
                snap.get("case_id", ""),
                snap.get("folder_name", ""),
                snap.get("time_key", ""),
                snap.get("npi_total"),
                snap.get("barthel_total"),
                snap.get("adl_dependence"),
                snap.get("mmse_total"),
                snap.get("cognitive_inverse"),
                snap.get("systolic_bp"),
                snap.get("heart_rate"),
                snap.get("spo2"),
                snap.get("excel_path", ""),
            ]
        )

    sheet_dict = workbook.create_sheet("标志物说明")
    sheet_dict.append(["编码", "名称", "方向", "临床说明"])
    for code, meta in BIOMARKER_DEFINITIONS.items():
        sheet_dict.append([code, meta.get("label", ""), meta.get("direction", ""), meta.get("clinical_note", "")])

    workbook.save(output_path)
    return output_path
